from __future__ import annotations

import csv
import uuid
import zipfile
from collections import Counter
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import cast

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.exceptions import ValidationFailure
from app.models.entities import (
    AnswerVersion,
    ExportJob,
    Questionnaire,
    QuestionnaireRow,
    RetrievalRun,
    Upload,
)
from app.models.enums import (
    AnswerStatus,
    ExecutionRunKind,
    ExportMode,
    ExportStatus,
    QuestionnaireRowStatus,
    ReproducibilityMode,
    SourceManifestKind,
    UploadKind,
)
from app.services.export_selection import (
    APPROVED_ANSWER_SELECTION_KIND,
    STATUS_PLACEHOLDER_SELECTION_KIND,
    ExportSelectionKind,
    approved_only_placeholder_text,
    latest_available_placeholder_text,
)
from app.services.reproducibility import (
    assert_execution_run_consistency,
    export_input_manifest,
    finish_execution_run,
    get_or_create_source_manifest,
    start_repro_run,
)
from app.services.storage import LocalObjectStorage
from app.services.workbooks import EXPECTED_HEADERS


@dataclass(frozen=True, slots=True)
class ExportRowSelection:
    selection_kind: ExportSelectionKind
    review_status: QuestionnaireRowStatus
    answer_version: AnswerVersion | None = None
    placeholder_text: str | None = None
    uses_unapproved_draft: bool = False

    def __post_init__(self) -> None:
        if self.selection_kind == APPROVED_ANSWER_SELECTION_KIND:
            if self.answer_version is None or self.placeholder_text is not None:
                raise ValueError(
                    "approved_answer selections require an answer version and no placeholder."
                )
            return
        if self.answer_version is not None or self.placeholder_text is None:
            raise ValueError(
                "status_placeholder selections require placeholder text and no answer version."
            )


@dataclass(frozen=True, slots=True)
class ResolvedExportRow:
    row: QuestionnaireRow
    selection: ExportRowSelection
    export_text: str


def _validated_approved_answer(
    session: Session,
    *,
    row: QuestionnaireRow,
) -> AnswerVersion | None:
    if row.approved_answer_version_id is None:
        return None
    approved = session.get(AnswerVersion, row.approved_answer_version_id)
    if approved is None:
        raise ValidationFailure(
            f"Row {row.source_row_id} references missing approved answer version {row.approved_answer_version_id}."
        )
    if approved.questionnaire_row_id != row.id or approved.case_id != row.case_id:
        raise ValidationFailure(
            f"Row {row.source_row_id} references an approved answer from another row/case."
        )
    return approved


def _selection_record(*, selection: ExportRowSelection) -> dict[str, object]:
    answer_version = selection.answer_version
    return {
        "selection_kind": selection.selection_kind,
        "review_status": selection.review_status.value,
        "answer_version_id": str(answer_version.id) if answer_version is not None else None,
        "answer_status": answer_version.status.value if answer_version is not None else None,
        "placeholder_text": selection.placeholder_text,
    }


def _latest_placeholder_selection(row: QuestionnaireRow) -> ExportRowSelection:
    return ExportRowSelection(
        selection_kind=STATUS_PLACEHOLDER_SELECTION_KIND,
        review_status=row.review_status,
        placeholder_text=latest_available_placeholder_text(row.review_status),
    )


def _resolve_export_selection(
    session: Session,
    *,
    row: QuestionnaireRow,
    mode: ExportMode,
) -> ExportRowSelection:
    approved = _validated_approved_answer(session, row=row)
    if mode == ExportMode.APPROVED_ONLY:
        if approved is None:
            if row.review_status == QuestionnaireRowStatus.APPROVED:
                raise ValidationFailure(
                    f"Row {row.source_row_id} is marked approved but has no approved answer for approved_only export."
                )
            return ExportRowSelection(
                selection_kind=STATUS_PLACEHOLDER_SELECTION_KIND,
                review_status=row.review_status,
                placeholder_text=approved_only_placeholder_text(row.review_status),
            )
        return ExportRowSelection(
            selection_kind=APPROVED_ANSWER_SELECTION_KIND,
            review_status=row.review_status,
            answer_version=approved,
        )

    if row.review_status in {
        QuestionnaireRowStatus.REJECTED,
        QuestionnaireRowStatus.FAILED,
        QuestionnaireRowStatus.NOT_STARTED,
        QuestionnaireRowStatus.SKIPPED,
    }:
        return _latest_placeholder_selection(row)

    latest = session.scalar(
        select(AnswerVersion)
        .where(AnswerVersion.questionnaire_row_id == row.id)
        .order_by(AnswerVersion.version_number.desc())
    )
    if latest is None:
        if row.review_status in {
            QuestionnaireRowStatus.NEEDS_REVIEW,
            QuestionnaireRowStatus.RUNNING,
        }:
            return _latest_placeholder_selection(row)
        raise ValidationFailure(
            f"Row {row.source_row_id} has no generated answer for latest_available export."
        )
    includes_unapproved_draft = (
        latest.status != AnswerStatus.ACCEPTED
        or row.approved_answer_version_id != latest.id
    )
    return ExportRowSelection(
        selection_kind=APPROVED_ANSWER_SELECTION_KIND,
        review_status=row.review_status,
        answer_version=latest,
        uses_unapproved_draft=includes_unapproved_draft,
    )


def _assert_export_selection_repro_ready(
    session: Session,
    *,
    row: QuestionnaireRow,
    selection: ExportRowSelection,
) -> None:
    answer_version = selection.answer_version
    if answer_version is None:
        return
    retrieval_run = session.get(RetrievalRun, answer_version.retrieval_run_id)
    if (
        answer_version.execution_run_id is None
        or answer_version.model_invocation_id is None
        or retrieval_run is None
        or retrieval_run.execution_run_id is None
    ):
        raise ValidationFailure(
            f"Row {row.source_row_id} answer version {answer_version.id} is missing strict-eval lineage."
        )


def _render_xlsx_payload(
    *,
    workbook,
    resolved_rows: list[ResolvedExportRow],
) -> bytes:
    worksheet = workbook[resolved_rows[0].row.source_sheet_name] if resolved_rows else workbook.active
    for resolved in resolved_rows:
        worksheet.cell(row=resolved.row.source_row_number, column=3).value = resolved.export_text
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _render_csv_payload(*, resolved_rows: list[ResolvedExportRow]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(EXPECTED_HEADERS)
    for resolved in resolved_rows:
        writer.writerow(
            [
                resolved.row.context_raw,
                resolved.row.question_raw,
                resolved.export_text,
            ]
        )
    return output.getvalue().encode("utf-8")


def _render_zip_payload(
    *,
    xlsx_name: str,
    xlsx_payload: bytes,
    csv_name: str,
    csv_payload: bytes,
) -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(xlsx_name, xlsx_payload)
        archive.writestr(csv_name, csv_payload)
    return output.getvalue()


def export_questionnaire(
    session: Session,
    *,
    storage: LocalObjectStorage,
    settings=None,
    questionnaire: Questionnaire,
    upload: Upload,
    mode: ExportMode,
    user_id=None,
    reproducibility_mode: ReproducibilityMode = ReproducibilityMode.BEST_EFFORT,
) -> ExportJob:
    payload = storage.read_bytes(upload.object_key)
    workbook = load_workbook(BytesIO(payload))
    worksheet = workbook[questionnaire.source_sheet_name]
    headers = tuple(
        (worksheet.cell(row=1, column=index).value or "").strip() for index in range(1, 4)
    )
    if headers != EXPECTED_HEADERS:
        raise ValidationFailure(
            f"Questionnaire export cannot prove source mapping because headers changed from {EXPECTED_HEADERS} to {headers}."
        )
    rows = session.scalars(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.questionnaire_id == questionnaire.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    ).all()
    resolved_rows: list[ResolvedExportRow] = []
    mapping: dict[str, object] = {}
    includes_unapproved_drafts = False
    placeholder_row_count = 0
    placeholder_status_counts: Counter[str] = Counter()
    row_selection: list[dict[str, object]] = []
    for row in rows:
        if row.source_sheet_name != questionnaire.source_sheet_name:
            raise ValidationFailure(
                f"Questionnaire row {row.id} references sheet {row.source_sheet_name}, expected {questionnaire.source_sheet_name}."
            )
        selection = _resolve_export_selection(
            session,
            row=row,
            mode=mode,
        )
        if reproducibility_mode == ReproducibilityMode.STRICT_EVAL:
            _assert_export_selection_repro_ready(
                session,
                row=row,
                selection=selection,
            )
        derived_row_id = (
            f"{questionnaire.source_file_name}:{row.source_sheet_name}:{row.source_row_number}"
        )
        if derived_row_id != row.source_row_id:
            raise ValidationFailure(
                f"Deterministic row mapping failed for row {row.id}; stored source_row_id {row.source_row_id} does not match derived {derived_row_id}."
            )
        export_text = (
            selection.answer_version.answer_text
            if selection.answer_version is not None
            else cast(str, selection.placeholder_text)
        )
        resolved = ResolvedExportRow(
            row=row,
            selection=selection,
            export_text=export_text,
        )
        resolved_rows.append(resolved)
        record = _selection_record(selection=selection)
        mapping[row.source_row_id] = {
            "row_number": row.source_row_number,
            **record,
        }
        row_selection.append(
            {
                "source_row_id": row.source_row_id,
                **record,
            }
        )
        if selection.selection_kind == STATUS_PLACEHOLDER_SELECTION_KIND:
            placeholder_row_count += 1
            placeholder_status_counts[selection.review_status.value] += 1
        includes_unapproved_drafts = (
            includes_unapproved_drafts or selection.uses_unapproved_draft
        )
    source_manifest = get_or_create_source_manifest(
        session,
        tenant_id=questionnaire.tenant_id,
        case_id=questionnaire.case_id,
        kind=SourceManifestKind.EXPORT_INPUT,
        manifest_json=export_input_manifest(
            questionnaire_id=questionnaire.id,
            source_upload_id=upload.id,
            source_upload_hash=upload.file_hash,
            export_mode=mode.value,
            row_selection=row_selection,
        ),
    )
    repro = start_repro_run(
        session,
        storage=None,
        settings=settings or get_settings(),
        kind=ExecutionRunKind.EXPORT,
        mode=reproducibility_mode,
        tenant_id=questionnaire.tenant_id,
        case_id=questionnaire.case_id,
        user_id=user_id,
        source_manifest=source_manifest,
        inputs_json={"questionnaire_id": str(questionnaire.id), "export_mode": mode.value},
    )

    xlsx_payload = _render_xlsx_payload(workbook=workbook, resolved_rows=resolved_rows)
    csv_payload = _render_csv_payload(resolved_rows=resolved_rows)
    base_name = questionnaire.source_file_name.removesuffix(".xlsx")
    xlsx_file_name = f"{base_name}_filled.xlsx"
    csv_file_name = f"{base_name}_filled.csv"
    zip_payload = _render_zip_payload(
        xlsx_name=xlsx_file_name,
        xlsx_payload=xlsx_payload,
        csv_name=csv_file_name,
        csv_payload=csv_payload,
    )

    export_object_key = f"exports/{questionnaire.case_id}/{questionnaire.id}/{uuid.uuid4()}.xlsx"
    exported = storage.save_bytes(
        object_key=export_object_key,
        payload=xlsx_payload,
    )
    csv_object_key = f"exports/{questionnaire.case_id}/{questionnaire.id}/{uuid.uuid4()}.csv"
    exported_csv = storage.save_bytes(
        object_key=csv_object_key,
        payload=csv_payload,
    )
    zip_object_key = f"exports/{questionnaire.case_id}/{questionnaire.id}/{uuid.uuid4()}.zip"
    exported_zip = storage.save_bytes(
        object_key=zip_object_key,
        payload=zip_payload,
    )

    output_upload = Upload(
        tenant_id=questionnaire.tenant_id,
        case_id=questionnaire.case_id,
        kind=UploadKind.EXPORT_XLSX,
        original_file_name=xlsx_file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        object_key=exported.object_key,
        file_hash=exported.file_hash,
        size_bytes=exported.size_bytes,
        payload=exported.payload,
    )
    csv_upload = Upload(
        tenant_id=questionnaire.tenant_id,
        case_id=questionnaire.case_id,
        kind=UploadKind.EXPORT_CSV,
        original_file_name=csv_file_name,
        media_type="text/csv",
        object_key=exported_csv.object_key,
        file_hash=exported_csv.file_hash,
        size_bytes=exported_csv.size_bytes,
        payload=exported_csv.payload,
    )
    zip_upload = Upload(
        tenant_id=questionnaire.tenant_id,
        case_id=questionnaire.case_id,
        kind=UploadKind.EXPORT_ZIP,
        original_file_name=f"{base_name}_filled.zip",
        media_type="application/zip",
        object_key=exported_zip.object_key,
        file_hash=exported_zip.file_hash,
        size_bytes=exported_zip.size_bytes,
        payload=exported_zip.payload,
    )
    session.add(output_upload)
    session.add(csv_upload)
    session.add(zip_upload)
    session.flush()
    export_job = ExportJob(
        tenant_id=questionnaire.tenant_id,
        case_id=questionnaire.case_id,
        questionnaire_id=questionnaire.id,
        source_upload_id=upload.id,
        output_upload_id=output_upload.id,
        execution_run_id=repro.execution_run.id,
        export_mode=mode,
        status=ExportStatus.COMPLETED,
        row_mapping_json=mapping,
        metadata_json={
            "includes_unapproved_drafts": includes_unapproved_drafts,
            "placeholder_row_count": placeholder_row_count,
            "placeholder_status_counts": dict(placeholder_status_counts),
            "csv_upload_id": str(csv_upload.id),
            "zip_upload_id": str(zip_upload.id),
        },
        error_detail=None,
    )
    session.add(export_job)
    session.flush()
    finish_execution_run(
        repro.execution_run,
        outputs_json={
            "export_job_id": str(export_job.id),
            "output_upload_id": str(output_upload.id),
            "csv_upload_id": str(csv_upload.id),
            "zip_upload_id": str(zip_upload.id),
            "includes_unapproved_drafts": includes_unapproved_drafts,
            "placeholder_row_count": placeholder_row_count,
        },
    )
    if reproducibility_mode == ReproducibilityMode.STRICT_EVAL:
        assert_execution_run_consistency(session, run=repro.execution_run)
    return export_job
