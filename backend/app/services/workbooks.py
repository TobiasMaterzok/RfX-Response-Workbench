from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook  # type: ignore[import-untyped]

from app.exceptions import ValidationFailure
from app.services.hashing import sha256_hex

EXPECTED_HEADERS = ("Context", "Question", "Answer")
HISTORICAL_SCHEMA_VERSION = "historical_qa_workbook.v1"
QUESTIONNAIRE_SCHEMA_VERSION = "questionnaire_workbook.v1"


@dataclass(frozen=True)
class ParsedWorkbookRow:
    source_row_number: int
    source_row_id: str
    context: str
    question: str
    answer: str


@dataclass(frozen=True)
class ParsedWorkbook:
    source_file_name: str
    source_sheet_name: str
    schema_version: str
    rows: list[ParsedWorkbookRow]
    file_hash: str


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        return str(value)
    return value.strip()


def infer_language(text: str) -> tuple[str, float]:
    lowered = text.lower()
    if any(token in lowered for token in (" welche ", " wie ", " und ", " der ", " die ", " das ")):
        return ("de", 0.7)
    return ("en", 0.6)


def parse_workbook_bytes(
    payload: bytes,
    *,
    source_file_name: str,
    schema_version: str,
    allow_empty_answer: bool,
) -> ParsedWorkbook:
    from io import BytesIO

    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    if not workbook.sheetnames:
        raise ValidationFailure(f"Workbook {source_file_name} has no worksheets.")
    source_sheet_name = workbook.sheetnames[0]
    worksheet = workbook[source_sheet_name]

    headers = tuple(
        _normalize_cell(worksheet.cell(row=1, column=index).value) for index in range(1, 4)
    )
    if headers != EXPECTED_HEADERS:
        raise ValidationFailure(
            f"Workbook {source_file_name} sheet {source_sheet_name} must have exact headers "
            f"{EXPECTED_HEADERS}, observed {headers}."
        )

    rows: list[ParsedWorkbookRow] = []
    for row_number in range(2, worksheet.max_row + 1):
        context = _normalize_cell(worksheet.cell(row=row_number, column=1).value)
        question = _normalize_cell(worksheet.cell(row=row_number, column=2).value)
        answer = _normalize_cell(worksheet.cell(row=row_number, column=3).value)
        if not question:
            raise ValidationFailure(
                f"Workbook {source_file_name} sheet {source_sheet_name} row {row_number} has empty Question."
            )
        if not allow_empty_answer and not answer:
            raise ValidationFailure(
                f"Workbook {source_file_name} sheet {source_sheet_name} row {row_number} has empty Answer."
            )
        if not context:
            raise ValidationFailure(
                f"Workbook {source_file_name} sheet {source_sheet_name} row {row_number} has empty Context."
            )
        source_row_id = f"{source_file_name}:{source_sheet_name}:{row_number}"
        rows.append(
            ParsedWorkbookRow(
                source_row_number=row_number,
                source_row_id=source_row_id,
                context=context,
                question=question,
                answer=answer,
            )
        )

    if not rows:
        raise ValidationFailure(
            f"Workbook {source_file_name} sheet {source_sheet_name} must contain at least one data row."
        )

    return ParsedWorkbook(
        source_file_name=source_file_name,
        source_sheet_name=source_sheet_name,
        schema_version=schema_version,
        rows=rows,
        file_hash=sha256_hex(payload),
    )
