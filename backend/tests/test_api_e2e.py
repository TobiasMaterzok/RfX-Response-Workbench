from __future__ import annotations

import csv
import time
import zipfile
from io import BytesIO, StringIO
from pathlib import Path
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select, text

from app.models.entities import BulkFillRowExecution, Questionnaire, RfxCase
from app.models.enums import BulkFillRowStatus, BulkFillStatus
from app.services.bulk_fill import create_initial_bulk_fill_request, run_bulk_fill_worker_once
from tests.seed_paths import historical_customer_dir


def _create_case_via_api(
    client,
    *,
    auth_headers: dict[str, str],
    repo_root: Path,
    name: str,
) -> dict[str, object]:
    base = historical_customer_dir(repo_root, "nordtransit_logistik_ag")
    with (
        (base / "nordtransit_logistik_ag_context_brief.pdf").open("rb") as pdf_file,
        (base / "nordtransit_logistik_ag_qa.xlsx").open("rb") as workbook_file,
    ):
        response = client.post(
            "/api/cases",
            headers=auth_headers,
            data={"name": name, "client_name": "NordTransit Logistik AG"},
            files={
                "pdf": ("nordtransit_logistik_ag_context_brief.pdf", pdf_file, "application/pdf"),
                "questionnaire": (
                    "nordtransit_logistik_ag_qa.xlsx",
                    workbook_file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
    assert response.status_code == 200
    return response.json()


def test_case_create_draft_export_flow(
    client, auth_headers: dict[str, str], repo_root: Path, container
) -> None:
    base = historical_customer_dir(repo_root, "nordtransit_logistik_ag")
    with (
        (base / "nordtransit_logistik_ag_context_brief.pdf").open("rb") as pdf_file,
        (base / "nordtransit_logistik_ag_qa.xlsx").open("rb") as workbook_file,
    ):
        response = client.post(
            "/api/cases",
            headers=auth_headers,
            data={"name": "NordTransit E2E", "client_name": "NordTransit Logistik AG"},
            files={
                "pdf": ("nordtransit_logistik_ag_context_brief.pdf", pdf_file, "application/pdf"),
                "questionnaire": (
                    "nordtransit_logistik_ag_qa.xlsx",
                    workbook_file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
    assert response.status_code == 200
    case_detail = response.json()
    assert case_detail["profile"]["schema_version"] == "rfx_case_profile.v3"
    assert len(case_detail["questionnaire_rows"]) == 30

    first_row = case_detail["questionnaire_rows"][0]
    second_row = case_detail["questionnaire_rows"][1]
    draft_response = client.post(
        f"/api/cases/{case_detail['id']}/rows/{first_row['id']}/draft",
        headers=auth_headers,
        json={"message": "Draft the answer with explicit grounding."},
    )
    assert draft_response.status_code == 200
    draft_body = draft_response.json()
    assert draft_body["thread"]["questionnaire_row_id"] == first_row["id"]
    assert draft_body["answer_version"]["generation_path"] == "two_stage_plan_render"
    assert draft_body["answer_version"]["llm_capture_stage"] == "answer_rendering"
    assert draft_body["retrieval"]["retrieval_action"] == "refresh_retrieval"
    assert draft_body["retrieval"]["sufficiency"]
    assert draft_body["evidence"]

    revise_response = client.post(
        f"/api/cases/{case_detail['id']}/rows/{first_row['id']}/draft",
        headers=auth_headers,
        json={
            "message": "ok but remove the bullet points. just a flowing text",
            "thread_id": draft_body["thread"]["id"],
            "revision_mode_override": "style_only",
        },
    )
    assert revise_response.status_code == 200
    revise_body = revise_response.json()
    assert revise_body["answer_version"]["generation_path"] == "render_only_reuse_plan"
    assert revise_body["answer_version"]["llm_capture_stage"] == "answer_rendering"
    assert (
        revise_body["answer_version"]["retrieval_run_id"]
        == draft_body["answer_version"]["retrieval_run_id"]
    )

    thread_response = client.get(
        f"/api/cases/{case_detail['id']}/threads/{draft_body['thread']['id']}",
        headers=auth_headers,
    )
    assert thread_response.status_code == 200
    thread_body = thread_response.json()
    assert [message["role"] for message in thread_body["messages"]] == [
        "user",
        "user",
        "assistant",
        "assistant",
    ]
    assert [
        message["answer_version_id"]
        for message in thread_body["messages"]
        if message["role"] == "user"
    ] == [None, None]
    assert {
        message["answer_version_id"]
        for message in thread_body["messages"]
        if message["role"] == "assistant"
    } == {
        draft_body["answer_version"]["id"],
        revise_body["answer_version"]["id"],
    }

    answers_response = client.get(
        f"/api/cases/{case_detail['id']}/rows/{first_row['id']}/answers",
        headers=auth_headers,
    )
    assert answers_response.status_code == 200
    assert answers_response.json()[0]["version_number"] == 2

    approve_response = client.post(
        f"/api/cases/{case_detail['id']}/rows/{first_row['id']}/approve",
        headers=auth_headers,
        json={"answer_version_id": answers_response.json()[0]["id"]},
    )
    assert approve_response.status_code == 200
    approve_body = approve_response.json()
    assert approve_body["review_status"] == "approved"
    assert approve_body["approved_answer_version_id"] == answers_response.json()[0]["id"]
    assert approve_body["approved_answer_text"] == answers_response.json()[0]["answer_text"]
    assert approve_body["current_answer"] == answers_response.json()[0]["answer_text"]
    assert approve_body["latest_attempt_state"] == "answer_available"
    assert approve_body["latest_attempt_thread_id"] == draft_body["thread"]["id"]

    bulk_fill_response = client.post(
        f"/api/cases/{case_detail['id']}/bulk-fill",
        headers=auth_headers,
        json={"note": "placeholder"},
    )
    assert bulk_fill_response.status_code == 200
    assert bulk_fill_response.json()["request"]["status"] in {
        "queued",
        "running",
        "completed",
        "completed_with_failures",
    }
    run_bulk_fill_worker_once(
        container,
        runner_id="api-e2e-worker",
        execution_mode="test_worker",
    )
    for _ in range(20):
        refreshed_case = client.get(
            f"/api/cases/{case_detail['id']}",
            headers=auth_headers,
        ).json()
        latest_bulk_fill = refreshed_case.get("latest_bulk_fill")
        if latest_bulk_fill and latest_bulk_fill["status"] in {
            "completed",
            "completed_with_failures",
            "failed",
            "cancelled",
        }:
            break
        time.sleep(0.05)

    approved_export_response = client.post(
        f"/api/cases/{case_detail['id']}/export",
        headers=auth_headers,
        json={"mode": "approved_only"},
    )
    assert approved_export_response.status_code == 200
    approved_export_body = approved_export_response.json()
    assert approved_export_body["export_mode"] == "approved_only"
    assert approved_export_body["includes_unapproved_drafts"] is False
    assert approved_export_body["placeholder_row_count"] == 29
    assert approved_export_body["csv_download_upload_id"]
    assert approved_export_body["zip_download_upload_id"]
    approved_download_response = client.get(
        f"/api/cases/downloads/{approved_export_body['download_upload_id']}",
        headers=auth_headers,
    )
    assert approved_download_response.status_code == 200
    approved_workbook = load_workbook(BytesIO(approved_download_response.content))
    approved_sheet = approved_workbook["QA"]
    assert (
        approved_sheet.cell(row=first_row["source_row_number"], column=3).value
        == answers_response.json()[0]["answer_text"]
    )
    approved_placeholder = "No approved answer exported due to status: needs review."
    assert (
        approved_sheet.cell(row=second_row["source_row_number"], column=3).value
        == approved_placeholder
    )
    approved_csv_download_response = client.get(
        f"/api/cases/downloads/{approved_export_body['csv_download_upload_id']}",
        headers=auth_headers,
    )
    assert approved_csv_download_response.status_code == 200
    assert approved_csv_download_response.headers["content-disposition"].endswith(
        '_filled.csv"'
    )
    approved_csv_rows = list(
        csv.reader(StringIO(approved_csv_download_response.content.decode("utf-8")))
    )
    assert approved_csv_rows[0] == ["Context", "Question", "Answer"]
    assert approved_csv_rows[1][2] == answers_response.json()[0]["answer_text"]
    assert approved_csv_rows[2][2] == approved_placeholder
    approved_zip_download_response = client.get(
        f"/api/cases/downloads/{approved_export_body['zip_download_upload_id']}",
        headers=auth_headers,
    )
    assert approved_zip_download_response.status_code == 200
    assert approved_zip_download_response.headers["content-disposition"].endswith(
        '_filled.zip"'
    )
    with zipfile.ZipFile(BytesIO(approved_zip_download_response.content)) as archive:
        assert sorted(archive.namelist()) == [
            "nordtransit_logistik_ag_qa_filled.csv",
            "nordtransit_logistik_ag_qa_filled.xlsx",
        ]

    second_answers_response = client.get(
        f"/api/cases/{case_detail['id']}/rows/{second_row['id']}/answers",
        headers=auth_headers,
    )
    assert second_answers_response.status_code == 200
    second_answers = second_answers_response.json()
    assert second_answers
    reject_response = client.post(
        f"/api/cases/{case_detail['id']}/rows/{second_row['id']}/reject",
        headers=auth_headers,
        json={"answer_version_id": second_answers[0]["id"]},
    )
    assert reject_response.status_code == 200
    reject_body = reject_response.json()
    assert reject_body["review_status"] == "rejected"
    assert reject_body["approved_answer_version_id"] is None
    assert reject_body["approved_answer_text"] is None
    assert reject_body["current_answer"] == second_answers[0]["answer_text"]
    assert reject_body["latest_attempt_state"] == "answer_available"
    assert reject_body["latest_attempt_thread_id"] is not None

    export_response = client.post(
        f"/api/cases/{case_detail['id']}/export",
        headers=auth_headers,
        json={"mode": "latest_available"},
    )
    assert export_response.status_code == 200
    export_body = export_response.json()
    assert export_body["export_mode"] == "latest_available"
    assert export_body["includes_unapproved_drafts"] is True
    assert export_body["placeholder_row_count"] >= 1
    assert export_body["csv_download_upload_id"]
    assert export_body["zip_download_upload_id"]
    download_response = client.get(
        f"/api/cases/downloads/{export_body['download_upload_id']}",
        headers=auth_headers,
    )
    assert download_response.status_code == 200
    workbook = load_workbook(BytesIO(download_response.content))
    assert "QA" in workbook.sheetnames
    latest_placeholder = "No latest answer exported due to status: rejected."
    assert (
        workbook["QA"].cell(row=second_row["source_row_number"], column=3).value
        == latest_placeholder
    )
    latest_csv_download_response = client.get(
        f"/api/cases/downloads/{export_body['csv_download_upload_id']}",
        headers=auth_headers,
    )
    assert latest_csv_download_response.status_code == 200
    assert latest_csv_download_response.headers["content-disposition"].endswith(
        '_filled.csv"'
    )
    latest_csv_rows = list(
        csv.reader(StringIO(latest_csv_download_response.content.decode("utf-8")))
    )
    assert latest_csv_rows[0] == ["Context", "Question", "Answer"]
    assert latest_csv_rows[2][2] == latest_placeholder
    latest_zip_download_response = client.get(
        f"/api/cases/downloads/{export_body['zip_download_upload_id']}",
        headers=auth_headers,
    )
    assert latest_zip_download_response.status_code == 200
    assert latest_zip_download_response.headers["content-disposition"].endswith(
        '_filled.zip"'
    )
    with zipfile.ZipFile(BytesIO(latest_zip_download_response.content)) as archive:
        assert sorted(archive.namelist()) == [
            "nordtransit_logistik_ag_qa_filled.csv",
            "nordtransit_logistik_ag_qa_filled.xlsx",
        ]
    assert download_response.headers["content-disposition"].endswith('_filled.xlsx"')

    cors_zip_download_response = client.get(
        f"/api/cases/downloads/{export_body['zip_download_upload_id']}",
        headers={**auth_headers, "Origin": "http://127.0.0.1:5173"},
    )
    assert cors_zip_download_response.status_code == 200
    assert (
        cors_zip_download_response.headers.get("access-control-expose-headers")
        == "Content-Disposition"
    )


def test_retry_failed_bulk_fill_route_returns_new_request(
    client,
    auth_headers: dict[str, str],
    session_factory,
    repo_root: Path,
) -> None:
    case_detail = _create_case_via_api(
        client,
        auth_headers=auth_headers,
        repo_root=repo_root,
        name="Retry Failed Route",
    )
    case_id = UUID(str(case_detail["id"]))
    with session_factory() as session:
        case = session.get(RfxCase, case_id)
        questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case_id))
        assert case is not None
        assert questionnaire is not None
        source_request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=case.created_by_user_id,
            note="retry-route",
        )
        row_run = session.scalar(
            select(BulkFillRowExecution).where(
                BulkFillRowExecution.bulk_fill_request_id == source_request.id
            )
        )
        assert row_run is not None
        source_request.status = BulkFillStatus.FAILED
        row_run.status = BulkFillRowStatus.FAILED
        session.commit()
        source_request_id = source_request.id

    response = client.post(
        f"/api/cases/{case_detail['id']}/bulk-fill/{source_request_id}/retry-failed",
        headers=auth_headers,
    )
    assert response.status_code == 200
    body = response.json()
    assert body["request"]["parent_request_id"] == str(source_request_id)
    assert body["request"]["status"] == "queued"


def test_resume_bulk_fill_route_returns_new_request(
    client,
    auth_headers: dict[str, str],
    session_factory,
    repo_root: Path,
) -> None:
    case_detail = _create_case_via_api(
        client,
        auth_headers=auth_headers,
        repo_root=repo_root,
        name="Resume Route",
    )
    case_id = UUID(str(case_detail["id"]))
    with session_factory() as session:
        case = session.get(RfxCase, case_id)
        questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case_id))
        assert case is not None
        assert questionnaire is not None
        source_request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=case.created_by_user_id,
            note="resume-route",
        )
        row_run = session.scalar(
            select(BulkFillRowExecution).where(
                BulkFillRowExecution.bulk_fill_request_id == source_request.id
            )
        )
        assert row_run is not None
        source_request.status = BulkFillStatus.CANCELLED
        row_run.status = BulkFillRowStatus.NOT_STARTED
        session.commit()
        source_request_id = source_request.id

    response = client.post(
        f"/api/cases/{case_detail['id']}/bulk-fill/{source_request_id}/resume",
        headers=auth_headers,
    )
    assert response.status_code == 200
    body = response.json()
    assert body["request"]["parent_request_id"] == str(source_request_id)
    assert body["request"]["status"] == "queued"


def test_cancel_bulk_fill_route_does_not_require_questionnaire_lookup(
    client,
    auth_headers: dict[str, str],
    session_factory,
    repo_root: Path,
) -> None:
    case_detail = _create_case_via_api(
        client,
        auth_headers=auth_headers,
        repo_root=repo_root,
        name="Cancel Route",
    )
    case_id = UUID(str(case_detail["id"]))
    with session_factory() as session:
        case = session.get(RfxCase, case_id)
        questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case_id))
        assert case is not None
        assert questionnaire is not None
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=case.created_by_user_id,
            note="cancel-route",
        )
        request_id = request.id
        session.execute(
            text("DELETE FROM questionnaires WHERE id = :questionnaire_id"),
            {"questionnaire_id": questionnaire.id.hex},
        )
        session.commit()

    response = client.post(
        f"/api/cases/{case_detail['id']}/bulk-fill/{request_id}/cancel",
        headers=auth_headers,
    )
    assert response.status_code == 200
    assert response.json()["request"]["status"] == "cancelled"
