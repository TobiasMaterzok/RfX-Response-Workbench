from __future__ import annotations

import csv
import json
import time
import zipfile
from datetime import timedelta
from io import BytesIO, StringIO
from pathlib import Path
from threading import Event, Thread
from uuid import UUID, uuid4

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session, sessionmaker

from app.config import Settings
from app.exceptions import ValidationFailure
from app.models.entities import (
    AnswerVersion,
    BulkFillRequest,
    BulkFillRowExecution,
    ChatMessage,
    ChatThread,
    ExecutionRun,
    Questionnaire,
    QuestionnaireRow,
    RfxCase,
    Upload,
)
from app.models.enums import (
    BulkFillEventType,
    BulkFillRowStatus,
    BulkFillStatus,
    ExecutionRunKind,
    ExecutionRunStatus,
    ExportMode,
    MessageRole,
    QuestionnaireRowStatus,
)
from app.services.ai import AnswerRenderGenerationResult, OpenAIAIService, StubAIService
from app.services.answers import draft_answer_for_row
from app.services.bulk_fill import (
    _require_request_scope,
    approve_answer_version,
    cancel_bulk_fill_request,
    claim_next_bulk_fill_request,
    create_initial_bulk_fill_request,
    detect_orphaned_bulk_fill_requests,
    execute_bulk_fill_request,
    list_bulk_fill_request_events,
    reject_row_answer,
    resume_bulk_fill_request,
    retry_failed_bulk_fill_request,
)
from app.services.cases import create_case_from_uploads, require_case_scope
from app.services.container import ServiceContainer
from app.services.exports import export_questionnaire
from app.services.identity import ensure_local_identity
from app.services.storage import LocalObjectStorage
from tests.seed_paths import historical_customer_dir


def sample_pdf_bytes(repo_root: Path) -> bytes:
    return (
        historical_customer_dir(repo_root, "nordtransit_logistik_ag")
        / "nordtransit_logistik_ag_context_brief.pdf"
    ).read_bytes()


def build_questionnaire_payload(rows: list[tuple[str, str, str]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "QA"
    worksheet["A1"] = "Context"
    worksheet["B1"] = "Question"
    worksheet["C1"] = "Answer"
    for index, (context, question, answer) in enumerate(rows, start=2):
        worksheet.cell(row=index, column=1).value = context
        worksheet.cell(row=index, column=2).value = question
        worksheet.cell(row=index, column=3).value = answer
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_csv_payload(payload: bytes) -> list[list[str]]:
    return list(csv.reader(StringIO(payload.decode("utf-8"))))


def read_zip_entries(payload: bytes) -> dict[str, bytes]:
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def row_for_case_number(
    session: Session,
    *,
    case_id,
    source_row_number: int,
) -> QuestionnaireRow:
    row = session.scalar(
        select(QuestionnaireRow)
        .where(
            QuestionnaireRow.case_id == case_id,
            QuestionnaireRow.source_row_number == source_row_number,
        )
    )
    assert row is not None
    return row


def create_case_with_rows(
    session: Session,
    *,
    tenant_id,
    user_id,
    container,
    repo_root: Path,
    ai_service: StubAIService,
    rows: list[tuple[str, str, str]],
    name: str = "Bulk Fill Case",
) -> RfxCase:
    return create_case_from_uploads(
        session,
        storage=container.storage,
        ai_service=ai_service,
        tenant_id=tenant_id,
        user_id=user_id,
        case_name=name,
        client_name="Eval Client",
        pdf_file_name="context.pdf",
        pdf_media_type="application/pdf",
        pdf_payload=sample_pdf_bytes(repo_root),
        questionnaire_file_name="qa.xlsx",
        questionnaire_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        questionnaire_payload=build_questionnaire_payload(rows),
    )


def questionnaire_for_case(session: Session, *, case_id) -> Questionnaire:
    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case_id))
    assert questionnaire is not None
    return questionnaire


def row_runs_for_request(session: Session, *, request_id) -> list[BulkFillRowExecution]:
    return session.scalars(
        select(BulkFillRowExecution)
        .where(BulkFillRowExecution.bulk_fill_request_id == request_id)
        .order_by(BulkFillRowExecution.created_at.asc())
    ).all()


def wait_for_request_status(
    session_factory: sessionmaker[Session],
    *,
    request_id,
    expected_status: str,
    timeout_seconds: float = 3.0,
) -> BulkFillRequest:
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        with session_factory() as session:
            request = session.get(BulkFillRequest, request_id)
            if request is not None and request.status.value == expected_status:
                return request
        time.sleep(0.05)
    raise AssertionError(f"Bulk-fill request {request_id} did not reach status {expected_status}.")


class FailingDraftAIService(StubAIService):
    def __init__(self, *, failing_question: str) -> None:
        self._failing_question = failing_question

    def render_answer(self, **kwargs):  # type: ignore[override]
        if kwargs["row_question"] == self._failing_question:
            raise ValidationFailure(f"Draft generation failed for row {kwargs['row_question']}.")
        return super().render_answer(**kwargs)


class SlowDraftAIService(StubAIService):
    def __init__(self) -> None:
        self.started = Event()
        self.release = Event()

    def render_answer(
        self,
        *,
        row_context: str,
        row_question: str,
        user_request: str,
        thread_history,
        answer_plan,
        output_mode: str,
        target_language: str,
        pipeline,
    ) -> AnswerRenderGenerationResult:
        self.started.set()
        self.release.wait(timeout=2)
        return super().render_answer(
            row_context=row_context,
            row_question=row_question,
            user_request=user_request,
            thread_history=thread_history,
            answer_plan=answer_plan,
            output_mode=output_mode,
            target_language=target_language,
            pipeline=pipeline,
        )


class FailingEmbeddingAIService(StubAIService):
    def embed_text(self, text: str, *, model_id: str | None = None) -> list[float]:
        raise ValidationFailure("Embedding generation failed for testing.")


def test_bulk_fill_happy_path_persists_answer_versions_and_snapshots(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[
                ("Context A", "Question A", ""),
                ("Context B", "Question B", ""),
                ("Context C", "Question C", ""),
            ],
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="bulk happy path",
        )
        request_id = request.id
        session.commit()

    execute_bulk_fill_request(container, request_id=request_id)

    with session_factory() as session:
        request = session.get(BulkFillRequest, request_id)
        assert request is not None
        assert request.status == BulkFillStatus.COMPLETED
        row_runs = row_runs_for_request(session, request_id=request_id)
        assert all(row_run.status == BulkFillRowStatus.DRAFTED for row_run in row_runs)
        answer_count = session.scalar(
            select(func.count())
            .select_from(AnswerVersion)
            .where(AnswerVersion.case_id == request.case_id)
        )
        assert answer_count == 3


def test_bulk_fill_render_prompt_omits_prior_thread_history(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[
                ("Context A", "Question A", ""),
            ],
        )
        row = session.scalar(
            select(QuestionnaireRow)
            .where(QuestionnaireRow.case_id == case.id)
            .order_by(QuestionnaireRow.source_row_number.asc())
        )
        assert row is not None
        first = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="bulk without thread history",
        )
        request_id = request.id
        session.commit()

    execute_bulk_fill_request(container, request_id=request_id)

    with session_factory() as session:
        latest_answer = session.scalar(
            select(AnswerVersion)
            .where(AnswerVersion.questionnaire_row_id == row.id)
            .order_by(AnswerVersion.version_number.desc())
        )
        assert latest_answer is not None
        assert latest_answer.version_number == 2
        assert latest_answer.chat_thread_id != first.thread.id
        prompt = latest_answer.llm_request_text
        assert prompt is not None
        assert "Draft the answer." not in prompt
        assert json.dumps(first.answer_version.answer_text, ensure_ascii=True)[1:-1] not in prompt
        rows = session.scalars(
            select(QuestionnaireRow).where(QuestionnaireRow.case_id == request.case_id)
        ).all()
        assert all(row.review_status == QuestionnaireRowStatus.NEEDS_REVIEW for row in rows)


def test_initial_bulk_fill_excludes_approved_rows(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[
                ("Context A", "Question A", ""),
                ("Context B", "Question B", ""),
            ],
            name="Approved Skip Case",
        )
        rows = session.scalars(
            select(QuestionnaireRow)
            .where(QuestionnaireRow.case_id == case.id)
            .order_by(QuestionnaireRow.source_row_number.asc())
        ).all()
        first = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=rows[0],
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        approve_answer_version(
            session,
            case=case,
            row=rows[0],
            answer_version_id=first.answer_version.id,
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="exclude approved rows",
        )
        session.flush()
        row_runs = row_runs_for_request(session, request_id=request.id)
        assert len(row_runs) == 1
        assert row_runs[0].questionnaire_row_id == rows[1].id


def test_initial_bulk_fill_fails_when_all_rows_are_already_approved(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="All Approved Bulk Fill Case",
        )
        row = session.scalar(
            select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id)
        )
        assert row is not None
        first = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        approve_answer_version(
            session,
            case=case,
            row=row,
            answer_version_id=first.answer_version.id,
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        with pytest.raises(ValidationFailure, match="all rows are already approved"):
            create_initial_bulk_fill_request(
                session,
                case=case,
                questionnaire=questionnaire,
                user_id=context.user.id,
                note="all approved",
            )


def test_bulk_fill_rejected_rows_use_fresh_generate_threads(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Rejected Bulk Fill Case",
        )
        row = session.scalar(
            select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id)
        )
        assert row is not None
        first = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        reject_row_answer(
            session,
            case=case,
            row=row,
            answer_version_id=first.answer_version.id,
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="rejected rows regenerate fresh",
        )
        request_id = request.id
        session.commit()

    execute_bulk_fill_request(container, request_id=request_id)

    with session_factory() as session:
        latest = session.scalar(
            select(AnswerVersion)
            .where(AnswerVersion.questionnaire_row_id == row.id)
            .order_by(AnswerVersion.version_number.desc())
        )
        row = session.get(QuestionnaireRow, row.id)
        assert latest is not None and row is not None
        assert latest.version_number == 2
        assert latest.chat_thread_id != first.thread.id
        assert row.review_status == QuestionnaireRowStatus.NEEDS_REVIEW


def test_rows_approved_after_queueing_are_skipped(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Approved Mid Queue Case",
        )
        row = session.scalar(
            select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id)
        )
        assert row is not None
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="approved after queueing",
        )
        request_id = request.id
        session.commit()

    with session_factory() as session:
        row = session.scalar(
            select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id)
        )
        assert row is not None
        result = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        approve_answer_version(
            session,
            case=case,
            row=row,
            answer_version_id=result.answer_version.id,
        )
        approved_answer_id = result.answer_version.id
        session.commit()

    execute_bulk_fill_request(container, request_id=request_id)

    with session_factory() as session:
        request = session.get(BulkFillRequest, request_id)
        row_runs = row_runs_for_request(session, request_id=request_id)
        row = session.scalar(
            select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id)
        )
        answer_count = session.scalar(
            select(func.count())
            .select_from(AnswerVersion)
            .where(AnswerVersion.questionnaire_row_id == row.id)
        )
        assert request is not None and row is not None
        assert request.status == BulkFillStatus.COMPLETED
        assert len(row_runs) == 1
        assert row_runs[0].status == BulkFillRowStatus.SKIPPED
        assert row.approved_answer_version_id == approved_answer_id
        assert answer_count == 1


def test_worker_missing_openai_key_fails_request_without_row_failures(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    real_ai_container = ServiceContainer(
        settings=settings,
        session_factory=session_factory,
        storage=LocalObjectStorage(settings),
        ai_service=OpenAIAIService(settings),
    )
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=real_ai_container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Missing Key Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="missing openai key",
        )
        request_id = request.id
        session.commit()

    execute_bulk_fill_request(real_ai_container, request_id=request_id)

    with session_factory() as session:
        request = session.get(BulkFillRequest, request_id)
        row_runs = row_runs_for_request(session, request_id=request_id)
        events = list_bulk_fill_request_events(session, request_id=request_id)
        assert request is not None
        assert request.status == BulkFillStatus.FAILED
        assert all(row_run.status == BulkFillRowStatus.NOT_STARTED for row_run in row_runs)
        assert not any(event.event_type == BulkFillEventType.ROW_FAILED for event in events)


def test_bulk_fill_marks_row_attempt_and_retrieval_runs_failed_on_early_errors(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Execution Failure Cleanup Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="execution cleanup",
        )
        request_id = request.id
        session.commit()

    container.ai_service = FailingEmbeddingAIService()
    execute_bulk_fill_request(container, request_id=request_id)

    with session_factory() as session:
        row_run_record = row_runs_for_request(session, request_id=request_id)[0]
        assert row_run_record.execution_run_id is not None
        row_execution_run = session.get(ExecutionRun, row_run_record.execution_run_id)
        retrieval_execution_run = session.scalar(
            select(ExecutionRun)
            .where(
                ExecutionRun.case_id == row_run_record.case_id,
                ExecutionRun.kind == ExecutionRunKind.RETRIEVAL,
            )
            .order_by(ExecutionRun.created_at.desc())
        )
        assert row_execution_run is not None
        assert retrieval_execution_run is not None
        assert row_execution_run.status == ExecutionRunStatus.FAILED
        assert retrieval_execution_run.status == ExecutionRunStatus.FAILED


def test_bulk_fill_partial_failure_and_retry_failed_rows(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    failing_ai = FailingDraftAIService(failing_question="Question B")
    container.ai_service = failing_ai
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=failing_ai,
            rows=[
                ("Context A", "Question A", ""),
                ("Context B", "Question B", ""),
                ("Context C", "Question C", ""),
            ],
            name="Bulk Fail Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="partial failure",
        )
        request_id = request.id
        session.commit()

    execute_bulk_fill_request(container, request_id=request_id)

    with session_factory() as session:
        request = session.get(BulkFillRequest, request_id)
        assert request is not None
        assert request.status == BulkFillStatus.COMPLETED_WITH_FAILURES
        row_runs = row_runs_for_request(session, request_id=request.id)
        assert [row_run.status for row_run in row_runs].count(BulkFillRowStatus.FAILED) == 1
        failed_row = session.scalar(
            select(QuestionnaireRow).where(
                QuestionnaireRow.case_id == request.case_id,
                QuestionnaireRow.review_status == QuestionnaireRowStatus.FAILED,
            )
        )
        assert failed_row is not None
        questionnaire = questionnaire_for_case(session, case_id=request.case_id)
        source_request = request
        retry_request = retry_failed_bulk_fill_request(
            session,
            case=require_case_scope(session, case_id=request.case_id, tenant_id=request.tenant_id),
            questionnaire=questionnaire,
            source_request=source_request,
            user_id=context.user.id,
        )
        retry_request_id = retry_request.id
        session.commit()

    container.ai_service = StubAIService()
    execute_bulk_fill_request(container, request_id=retry_request_id)

    with session_factory() as session:
        retry_request = session.get(BulkFillRequest, retry_request_id)
        assert retry_request is not None
        assert retry_request.status == BulkFillStatus.COMPLETED
        retried_runs = row_runs_for_request(session, request_id=retry_request.id)
        assert len(retried_runs) == 1
        assert retried_runs[0].status == BulkFillRowStatus.DRAFTED
        retried_row_attempts = session.scalars(
            select(BulkFillRowExecution)
            .where(BulkFillRowExecution.questionnaire_row_id == retried_runs[0].questionnaire_row_id)
            .order_by(BulkFillRowExecution.attempt_number.asc())
        ).all()
        assert [attempt.attempt_number for attempt in retried_row_attempts] == [1, 2]


def test_resume_interrupted_job_does_not_overwrite_successful_rows(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[
                ("Context A", "Question A", ""),
                ("Context B", "Question B", ""),
            ],
            name="Resume Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="resume",
        )
        row_runs = row_runs_for_request(session, request_id=request.id)
        first_row = session.get(QuestionnaireRow, row_runs[0].questionnaire_row_id)
        assert first_row is not None
        result = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=first_row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        row_runs[0].status = BulkFillRowStatus.DRAFTED
        row_runs[0].answer_version_id = result.answer_version.id
        request.status = BulkFillStatus.CANCELLED
        session.commit()

        resume_request = resume_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            source_request=request,
            user_id=context.user.id,
        )
        resume_request_id = resume_request.id
        session.commit()

    execute_bulk_fill_request(container, request_id=resume_request_id)

    with session_factory() as session:
        first_row = session.scalar(
            select(QuestionnaireRow)
            .where(QuestionnaireRow.case_id == case.id)
            .order_by(QuestionnaireRow.source_row_number.asc())
        )
        assert first_row is not None
        first_row_version_count = session.scalar(
            select(func.count())
            .select_from(AnswerVersion)
            .where(AnswerVersion.questionnaire_row_id == first_row.id)
        )
        assert first_row_version_count == 1


def test_cancel_queued_request_marks_rows_explicitly(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", ""), ("Context B", "Question B", "")],
            name="Cancel Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="cancel queued",
        )
        request = cancel_bulk_fill_request(session, request=request)
        session.commit()
        assert request.status == BulkFillStatus.CANCELLED
        row_runs = row_runs_for_request(session, request_id=request.id)
        assert all(row_run.status == BulkFillRowStatus.CANCELLED for row_run in row_runs)


def test_cancel_running_request_is_explicit(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    slow_ai = SlowDraftAIService()
    container.ai_service = slow_ai
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=slow_ai,
            rows=[("Context A", "Question A", ""), ("Context B", "Question B", "")],
            name="Running Cancel Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="cancel running",
        )
        request_id = request.id
        session.commit()

    worker = Thread(
        target=execute_bulk_fill_request,
        kwargs={"container": container, "request_id": request_id},
        daemon=True,
    )
    worker.start()
    assert slow_ai.started.wait(timeout=2)
    with session_factory() as session:
        request = session.get(BulkFillRequest, request_id)
        assert request is not None
        cancel_bulk_fill_request(session, request=request)
        assert request.status == BulkFillStatus.CANCEL_REQUESTED
        session.commit()
    slow_ai.release.set()
    worker.join(timeout=2)
    request = wait_for_request_status(
        session_factory,
        request_id=request_id,
        expected_status="cancelled",
    )
    assert request.status == BulkFillStatus.CANCELLED


def test_worker_claim_path_and_duplicate_claim_prevention(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Claim Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="claim",
        )
        request_id = request.id
        session.commit()

    with session_factory() as session:
        claimed = claim_next_bulk_fill_request(
            session,
            runner_id="worker-1",
            execution_mode="worker_cli",
        )
        assert claimed is not None
        assert claimed.id == request_id
        assert claimed.status == BulkFillStatus.RUNNING
        assert claimed.runner_id == "worker-1"
        assert claimed.claim_id is not None
        session.commit()

    with session_factory() as session:
        duplicate = claim_next_bulk_fill_request(
            session,
            runner_id="worker-2",
            execution_mode="worker_cli",
        )
        assert duplicate is None


def test_stale_orphaned_claim_detection_is_explicit(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Orphan Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="orphan",
        )
        request_id = request.id
        session.commit()

    with session_factory() as session:
        claimed = claim_next_bulk_fill_request(
            session,
            runner_id="worker-1",
            execution_mode="worker_cli",
        )
        assert claimed is not None
        claimed.heartbeat_at = claimed.claimed_at - timedelta(minutes=20)  # type: ignore[operator]
        session.commit()

    with session_factory() as session:
        stale = detect_orphaned_bulk_fill_requests(session)
        assert stale and stale[0].id == request_id
        session.commit()
        orphaned = session.get(BulkFillRequest, request_id)
        assert orphaned is not None
        assert orphaned.status == BulkFillStatus.ORPHANED
        events = list_bulk_fill_request_events(session, request_id=request_id)
        assert any(event.event_type == BulkFillEventType.ORPHANED for event in events)


def test_impossible_retry_and_resume_states_fail_loudly(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Impossible Retry Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="impossible",
        )
        session.commit()
        execute_bulk_fill_request(container, request_id=request.id)
        completed = session.get(BulkFillRequest, request.id)
        assert completed is not None
        session.refresh(completed)
        assert completed.status == BulkFillStatus.COMPLETED
        with pytest.raises(ValidationFailure, match="cannot retry failed rows"):
            retry_failed_bulk_fill_request(
                session,
                case=case,
                questionnaire=questionnaire,
                source_request=completed,
                user_id=context.user.id,
            )
        with pytest.raises(ValidationFailure, match="cannot be resumed"):
            resume_bulk_fill_request(
                session,
                case=case,
                questionnaire=questionnaire,
                source_request=completed,
                user_id=context.user.id,
            )


def test_approve_and_reject_flow_is_explicit(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Approve Reject Case",
        )
        row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
        assert row is not None
        result = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        approve_answer_version(
            session,
            case=case,
            row=row,
            answer_version_id=result.answer_version.id,
        )
        session.flush()
        assert row.review_status == QuestionnaireRowStatus.APPROVED
        assert row.approved_answer_version_id == result.answer_version.id
        version = session.get(AnswerVersion, result.answer_version.id)
        assert version is not None
        assert version.status.value == "accepted"

        reject_row_answer(
            session,
            case=case,
            row=row,
            answer_version_id=result.answer_version.id,
        )
        session.flush()
        assert row.review_status == QuestionnaireRowStatus.REJECTED
        assert row.approved_answer_version_id is None

        second = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Regenerate the answer.",
            thread=result.thread,
        )
        session.flush()
        assert row.review_status == QuestionnaireRowStatus.NEEDS_REVIEW
        assert second.answer_version.version_number == 2


def test_invalid_approval_reference_fails_loudly(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", ""), ("Context B", "Question B", "")],
            name="Invalid Approve Case",
        )
        rows = session.scalars(
            select(QuestionnaireRow)
            .where(QuestionnaireRow.case_id == case.id)
            .order_by(QuestionnaireRow.source_row_number.asc())
        ).all()
        first = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=rows[0],
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        with pytest.raises(ValidationFailure, match="cannot be approved"):
            approve_answer_version(
                session,
                case=case,
                row=rows[1],
                answer_version_id=first.answer_version.id,
            )


def test_approved_only_export_writes_approved_answers_and_status_placeholders(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[
                ("Context A", "Question A", ""),
                ("Context B", "Question B", ""),
                ("Context C", "Question C", ""),
                ("Context D", "Question D", ""),
                ("Context E", "Question E", ""),
                ("Context F", "Question F", ""),
                ("Context G", "Question G", ""),
            ],
            name="Approved Export Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        upload = session.get(Upload, questionnaire.upload_id)
        assert upload is not None
        rows = session.scalars(
            select(QuestionnaireRow)
            .where(QuestionnaireRow.case_id == case.id)
            .order_by(QuestionnaireRow.source_row_number.asc())
        ).all()
        approved_result = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=rows[0],
            user_id=context.user.id,
            user_message="Draft the approved answer.",
            thread=None,
        )
        approve_answer_version(
            session,
            case=case,
            row=rows[0],
            answer_version_id=approved_result.answer_version.id,
        )
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=rows[2],
            user_id=context.user.id,
            user_message="Draft the needs-review answer.",
            thread=None,
        )
        rejected_result = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=rows[4],
            user_id=context.user.id,
            user_message="Draft the rejected answer.",
            thread=None,
        )
        reject_row_answer(
            session,
            case=case,
            row=rows[4],
            answer_version_id=rejected_result.answer_version.id,
        )
        rows[3].review_status = QuestionnaireRowStatus.RUNNING
        rows[5].review_status = QuestionnaireRowStatus.FAILED
        rows[6].review_status = QuestionnaireRowStatus.SKIPPED
        session.flush()

        export_job = export_questionnaire(
            session,
            storage=container.storage,
            questionnaire=questionnaire,
            upload=upload,
            mode=ExportMode.APPROVED_ONLY,
        )
        assert export_job.export_mode == ExportMode.APPROVED_ONLY
        assert export_job.metadata_json["includes_unapproved_drafts"] is False
        assert export_job.metadata_json["placeholder_row_count"] == 6
        assert export_job.metadata_json["placeholder_status_counts"] == {
            "not_started": 1,
            "needs_review": 1,
            "running": 1,
            "rejected": 1,
            "failed": 1,
            "skipped": 1,
        }
        approved_entry = export_job.row_mapping_json[rows[0].source_row_id]
        assert approved_entry["selection_kind"] == "approved_answer"
        assert approved_entry["review_status"] == "approved"
        assert approved_entry["answer_version_id"] == str(approved_result.answer_version.id)
        assert approved_entry["answer_status"] == "accepted"
        assert approved_entry["placeholder_text"] is None
        placeholder_entry = export_job.row_mapping_json[rows[2].source_row_id]
        assert placeholder_entry["selection_kind"] == "status_placeholder"
        assert placeholder_entry["review_status"] == "needs_review"
        assert placeholder_entry["answer_version_id"] is None
        assert placeholder_entry["answer_status"] is None
        assert (
            placeholder_entry["placeholder_text"]
            == "No approved answer exported due to status: needs review."
        )
        csv_upload_id = export_job.metadata_json.get("csv_upload_id")
        assert isinstance(csv_upload_id, str)
        zip_upload_id = export_job.metadata_json.get("zip_upload_id")
        assert isinstance(zip_upload_id, str)

        download_upload = session.get(Upload, export_job.output_upload_id)
        assert download_upload is not None
        workbook = load_workbook(BytesIO(container.storage.read_bytes(download_upload.object_key)))
        worksheet = workbook[questionnaire.source_sheet_name]
        csv_upload = session.get(Upload, UUID(csv_upload_id))
        assert csv_upload is not None
        assert csv_upload.kind.value == "export_csv"
        csv_rows = parse_csv_payload(container.storage.read_bytes(csv_upload.object_key))
        zip_upload = session.get(Upload, UUID(zip_upload_id))
        assert zip_upload is not None
        assert zip_upload.kind.value == "export_zip"
        zip_entries = read_zip_entries(container.storage.read_bytes(zip_upload.object_key))
        assert csv_rows[0] == ["Context", "Question", "Answer"]
        assert sorted(zip_entries) == ["qa_filled.csv", "qa_filled.xlsx"]
        assert zip_entries["qa_filled.csv"] == container.storage.read_bytes(csv_upload.object_key)
        assert (
            worksheet.cell(row=rows[0].source_row_number, column=3).value
            == approved_result.answer_version.answer_text
        )
        assert csv_rows[1] == [
            rows[0].context_raw,
            rows[0].question_raw,
            approved_result.answer_version.answer_text,
        ]
        assert (
            worksheet.cell(row=rows[1].source_row_number, column=3).value
            == "No approved answer exported due to status: not started."
        )
        assert csv_rows[2] == [
            rows[1].context_raw,
            rows[1].question_raw,
            "No approved answer exported due to status: not started.",
        ]
        assert (
            worksheet.cell(row=rows[2].source_row_number, column=3).value
            == "No approved answer exported due to status: needs review."
        )
        assert (
            worksheet.cell(row=rows[3].source_row_number, column=3).value
            == "No approved answer exported due to status: running."
        )
        assert (
            worksheet.cell(row=rows[4].source_row_number, column=3).value
            == "No approved answer exported due to status: rejected."
        )
        assert (
            worksheet.cell(row=rows[5].source_row_number, column=3).value
            == "No approved answer exported due to status: failed."
        )
        assert (
            worksheet.cell(row=rows[6].source_row_number, column=3).value
            == "No approved answer exported due to status: skipped."
        )


def test_latest_available_export_mode_is_explicit_and_deterministic(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[
                ("Context A", "Question A", ""),
                ("Context B", "Question B", ""),
                ("Context C", "Question C", ""),
                ("Context D", "Question D", ""),
                ("Context E", "Question E", ""),
                ("Context F", "Question F", ""),
            ],
            name="Latest Export Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        upload = session.get(Upload, questionnaire.upload_id)
        assert upload is not None
        rows = session.scalars(
            select(QuestionnaireRow)
            .where(QuestionnaireRow.case_id == case.id)
            .order_by(QuestionnaireRow.source_row_number.asc())
        ).all()
        first_result = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=rows[0],
            user_id=context.user.id,
            user_message="Draft the first answer.",
            thread=None,
        )
        running_result = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=rows[1],
            user_id=context.user.id,
            user_message="Draft the running answer.",
            thread=None,
        )
        rejected_result = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=rows[2],
            user_id=context.user.id,
            user_message="Draft the rejected answer.",
            thread=None,
        )
        reject_row_answer(
            session,
            case=case,
            row=rows[2],
            answer_version_id=rejected_result.answer_version.id,
        )
        rows[1].review_status = QuestionnaireRowStatus.RUNNING
        rows[3].review_status = QuestionnaireRowStatus.RUNNING
        rows[4].review_status = QuestionnaireRowStatus.FAILED
        rows[5].review_status = QuestionnaireRowStatus.SKIPPED
        session.flush()

        export_job = export_questionnaire(
            session,
            storage=container.storage,
            questionnaire=questionnaire,
            upload=upload,
            mode=ExportMode.LATEST_AVAILABLE,
        )
        assert export_job.export_mode == ExportMode.LATEST_AVAILABLE
        assert export_job.metadata_json["includes_unapproved_drafts"] is True
        assert export_job.metadata_json["placeholder_row_count"] == 4
        assert export_job.metadata_json["placeholder_status_counts"] == {
            "rejected": 1,
            "running": 1,
            "failed": 1,
            "skipped": 1,
        }
        csv_upload_id = export_job.metadata_json.get("csv_upload_id")
        assert isinstance(csv_upload_id, str)
        zip_upload_id = export_job.metadata_json.get("zip_upload_id")
        assert isinstance(zip_upload_id, str)
        assert set(export_job.row_mapping_json) == {row.source_row_id for row in rows}
        first_entry = export_job.row_mapping_json[rows[0].source_row_id]
        assert first_entry["selection_kind"] == "approved_answer"
        assert first_entry["placeholder_text"] is None
        running_entry = export_job.row_mapping_json[rows[1].source_row_id]
        assert running_entry["selection_kind"] == "approved_answer"
        assert running_entry["placeholder_text"] is None
        placeholder_entry = export_job.row_mapping_json[rows[2].source_row_id]
        assert placeholder_entry["selection_kind"] == "status_placeholder"
        assert placeholder_entry["review_status"] == "rejected"
        assert placeholder_entry["answer_version_id"] is None
        assert (
            placeholder_entry["placeholder_text"]
            == "No latest answer exported due to status: rejected."
        )

        download_upload = session.get(Upload, export_job.output_upload_id)
        assert download_upload is not None
        workbook = load_workbook(BytesIO(container.storage.read_bytes(download_upload.object_key)))
        worksheet = workbook[questionnaire.source_sheet_name]
        csv_upload = session.get(Upload, UUID(csv_upload_id))
        assert csv_upload is not None
        csv_rows = parse_csv_payload(container.storage.read_bytes(csv_upload.object_key))
        zip_upload = session.get(Upload, UUID(zip_upload_id))
        assert zip_upload is not None
        zip_entries = read_zip_entries(container.storage.read_bytes(zip_upload.object_key))
        assert csv_rows[0] == ["Context", "Question", "Answer"]
        assert sorted(zip_entries) == ["qa_filled.csv", "qa_filled.xlsx"]
        assert zip_entries["qa_filled.csv"] == container.storage.read_bytes(csv_upload.object_key)
        assert (
            worksheet.cell(row=rows[0].source_row_number, column=3).value
            == first_result.answer_version.answer_text
        )
        assert (
            worksheet.cell(row=rows[1].source_row_number, column=3).value
            == running_result.answer_version.answer_text
        )
        assert (
            worksheet.cell(row=rows[2].source_row_number, column=3).value
            == "No latest answer exported due to status: rejected."
        )
        assert (
            worksheet.cell(row=rows[3].source_row_number, column=3).value
            == "No latest answer exported due to status: running."
        )
        assert (
            worksheet.cell(row=rows[4].source_row_number, column=3).value
            == "No latest answer exported due to status: failed."
        )
        assert (
            worksheet.cell(row=rows[5].source_row_number, column=3).value
            == "No latest answer exported due to status: skipped."
        )
        assert csv_rows[1] == [
            rows[0].context_raw,
            rows[0].question_raw,
            first_result.answer_version.answer_text,
        ]
        assert csv_rows[2] == [
            rows[1].context_raw,
            rows[1].question_raw,
            running_result.answer_version.answer_text,
        ]
        assert csv_rows[3] == [
            rows[2].context_raw,
            rows[2].question_raw,
            "No latest answer exported due to status: rejected.",
        ]
        assert csv_rows[4] == [
            rows[3].context_raw,
            rows[3].question_raw,
            "No latest answer exported due to status: running.",
        ]
        assert csv_rows[5] == [
            rows[4].context_raw,
            rows[4].question_raw,
            "No latest answer exported due to status: failed.",
        ]
        assert csv_rows[6] == [
            rows[5].context_raw,
            rows[5].question_raw,
            "No latest answer exported due to status: skipped.",
        ]


def test_approved_only_export_prefers_explicit_approved_answer_over_newer_rejected_draft(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Approved Precedence Export Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        upload = session.get(Upload, questionnaire.upload_id)
        assert upload is not None
        row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
        assert row is not None
        approved_version = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the first answer.",
            thread=None,
        )
        approve_answer_version(
            session,
            case=case,
            row=row,
            answer_version_id=approved_version.answer_version.id,
        )
        rejected_version = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the second answer.",
            thread=None,
        )
        reject_row_answer(
            session,
            case=case,
            row=row,
            answer_version_id=rejected_version.answer_version.id,
        )
        session.flush()

        export_job = export_questionnaire(
            session,
            storage=container.storage,
            questionnaire=questionnaire,
            upload=upload,
            mode=ExportMode.APPROVED_ONLY,
        )
        assert export_job.metadata_json["placeholder_row_count"] == 0
        csv_upload_id = export_job.metadata_json.get("csv_upload_id")
        assert isinstance(csv_upload_id, str)
        zip_upload_id = export_job.metadata_json.get("zip_upload_id")
        assert isinstance(zip_upload_id, str)
        mapping_entry = export_job.row_mapping_json[row.source_row_id]
        assert mapping_entry["selection_kind"] == "approved_answer"
        assert mapping_entry["review_status"] == "rejected"
        assert mapping_entry["answer_version_id"] == str(approved_version.answer_version.id)
        assert mapping_entry["answer_status"] == "accepted"
        assert mapping_entry["placeholder_text"] is None

        download_upload = session.get(Upload, export_job.output_upload_id)
        assert download_upload is not None
        workbook = load_workbook(BytesIO(container.storage.read_bytes(download_upload.object_key)))
        worksheet = workbook[questionnaire.source_sheet_name]
        csv_upload = session.get(Upload, UUID(csv_upload_id))
        assert csv_upload is not None
        csv_rows = parse_csv_payload(container.storage.read_bytes(csv_upload.object_key))
        zip_upload = session.get(Upload, UUID(zip_upload_id))
        assert zip_upload is not None
        zip_entries = read_zip_entries(container.storage.read_bytes(zip_upload.object_key))
        assert (
            worksheet.cell(row=row.source_row_number, column=3).value
            == approved_version.answer_version.answer_text
        )
        assert csv_rows[1] == [
            row.context_raw,
            row.question_raw,
            approved_version.answer_version.answer_text,
        ]
        assert sorted(zip_entries) == ["qa_filled.csv", "qa_filled.xlsx"]


def test_approved_only_export_fails_on_corrupt_approved_state(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Corrupt Approved Export Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        upload = session.get(Upload, questionnaire.upload_id)
        assert upload is not None
        row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
        assert row is not None

        row.review_status = QuestionnaireRowStatus.APPROVED
        with pytest.raises(ValidationFailure, match="marked approved but has no approved answer"):
            export_questionnaire(
                session,
                storage=container.storage,
                questionnaire=questionnaire,
                upload=upload,
                mode=ExportMode.APPROVED_ONLY,
            )

        row.review_status = QuestionnaireRowStatus.NEEDS_REVIEW
        row.approved_answer_version_id = uuid4()
        with pytest.raises(ValidationFailure, match="references missing approved answer version"):
            export_questionnaire(
                session,
                storage=container.storage,
                questionnaire=questionnaire,
                upload=upload,
                mode=ExportMode.APPROVED_ONLY,
            )


def test_case_and_tenant_isolation_hold_for_bulk_fill_scope_checks(
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        local_context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=local_context.tenant.id,
            user_id=local_context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context A", "Question A", "")],
            name="Isolation Case",
        )
        questionnaire = questionnaire_for_case(session, case_id=case.id)
        request = create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=local_context.user.id,
            note="scope",
        )
        other_case = create_case_with_rows(
            session,
            tenant_id=local_context.tenant.id,
            user_id=local_context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[("Context B", "Question B", "")],
            name="Other Case",
        )
        session.flush()
        with pytest.raises(ValidationFailure, match="not available for case"):
            _require_request_scope(
                session,
                request_id=request.id,
                case=other_case,
            )


def test_case_detail_projects_latest_attempt_states_and_answer_thread_ids(
    client,
    auth_headers: dict[str, str],
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_rows(
            session,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            container=container,
            repo_root=repo_root,
            ai_service=StubAIService(),
            rows=[
                ("Context A", "Question A", ""),
                ("Context B", "Question B", ""),
                ("Context C", "Question C", ""),
            ],
            name="Attempt State Projection Case",
        )
        answer_row = row_for_case_number(session, case_id=case.id, source_row_number=3)
        pending_row = row_for_case_number(session, case_id=case.id, source_row_number=4)
        draft = draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=answer_row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
        )
        pending_thread = ChatThread(
            tenant_id=context.tenant.id,
            case_id=case.id,
            questionnaire_row_id=pending_row.id,
            created_by_user_id=context.user.id,
            title="Pending attempt",
        )
        session.add(pending_thread)
        session.flush()
        session.add(
            ChatMessage(
                tenant_id=context.tenant.id,
                case_id=case.id,
                questionnaire_row_id=pending_row.id,
                thread_id=pending_thread.id,
                role=MessageRole.USER,
                content="Draft the answer.",
            )
        )
        session.commit()

    response = client.get(f"/api/cases/{case.id}", headers=auth_headers)
    assert response.status_code == 200
    body = response.json()
    rows_by_number = {
        row["source_row_number"]: row for row in body["questionnaire_rows"]
    }
    assert rows_by_number[2]["latest_attempt_state"] == "none"
    assert rows_by_number[2]["latest_attempt_thread_id"] is None
    assert rows_by_number[3]["latest_attempt_state"] == "answer_available"
    assert rows_by_number[3]["latest_attempt_thread_id"] == str(draft.thread.id)
    assert rows_by_number[4]["latest_attempt_state"] == "pending_no_answer"
    assert rows_by_number[4]["latest_attempt_thread_id"] == str(pending_thread.id)

    answers_response = client.get(
        f"/api/cases/{case.id}/rows/{answer_row.id}/answers",
        headers=auth_headers,
    )
    assert answers_response.status_code == 200
    answers_body = answers_response.json()
    assert answers_body[0]["chat_thread_id"] == str(draft.thread.id)


def test_thread_route_returns_failed_no_answer_thread_payload(
    client,
    auth_headers: dict[str, str],
    session_factory,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    container.ai_service = FailingDraftAIService(failing_question="Question B")
    try:
        with session_factory() as session:
            context = ensure_local_identity(session, settings)
            case = create_case_with_rows(
                session,
                tenant_id=context.tenant.id,
                user_id=context.user.id,
                container=container,
                repo_root=repo_root,
                ai_service=StubAIService(),
                rows=[
                    ("Context A", "Question A", ""),
                    ("Context B", "Question B", ""),
                ],
                name="Failed Thread Route Case",
            )
            questionnaire = questionnaire_for_case(session, case_id=case.id)
            request = create_initial_bulk_fill_request(
                session,
                case=case,
                questionnaire=questionnaire,
                user_id=context.user.id,
                note="failed thread route",
            )
            request_id = request.id
            session.commit()

        execute_bulk_fill_request(container, request_id=request_id)

        case_response = client.get(f"/api/cases/{case.id}", headers=auth_headers)
        assert case_response.status_code == 200
        case_body = case_response.json()
        failed_row = next(
            row
            for row in case_body["questionnaire_rows"]
            if row["source_row_number"] == 3
        )
        assert failed_row["latest_attempt_state"] == "failed_no_answer"
        assert failed_row["latest_attempt_thread_id"] is not None

        thread_response = client.get(
            f"/api/cases/{case.id}/threads/{failed_row['latest_attempt_thread_id']}",
            headers=auth_headers,
        )
        assert thread_response.status_code == 200
        thread_body = thread_response.json()
        assert thread_body["thread_state"] == "failed_no_answer"
        assert thread_body["answer_version"] is None
        assert thread_body["retrieval"] is not None
        assert thread_body["evidence"]
        assert thread_body["messages"]
        assert "Draft generation failed for row Question B." in thread_body["failure_detail"]
    finally:
        container.ai_service = StubAIService()
