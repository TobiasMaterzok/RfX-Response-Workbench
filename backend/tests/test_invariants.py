from __future__ import annotations

import json
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy import text as sql_text
from sqlalchemy.orm import Session

from app.config import Settings, build_settings
from app.exceptions import ScopeViolation, ValidationFailure
from app.models.entities import (
    AnswerVersion,
    CaseProfile,
    CaseProfileItem,
    ChatMessage,
    ChatThread,
    ExecutionRun,
    HistoricalCaseProfile,
    HistoricalCaseProfileItem,
    HistoricalClientPackage,
    HistoricalDataset,
    HistoricalQARow,
    HistoricalWorkbook,
    ModelInvocation,
    PdfChunk,
    PdfPage,
    Questionnaire,
    QuestionnaireRow,
    RetrievalRun,
    RfxCase,
    Tenant,
    Upload,
    User,
)
from app.models.enums import (
    ApprovalStatus,
    ExportMode,
    LLMCaptureStatus,
    MembershipRole,
    MessageRole,
    ModelInvocationKind,
    ReproducibilityMode,
    UploadKind,
)
from app.pipeline.config import (
    artifact_index_hashes,
    case_profile_index_payload,
    historical_index_payload,
    resolve_pipeline_selection,
)
from app.schemas.case_profile import CaseProfileDocument
from app.services.ai import StubAIService
from app.services.answers import draft_answer_for_row
from app.services.case_profiles import persist_case_profile
from app.services.cases import create_case_from_uploads
from app.services.hashing import sha256_text
from app.services.identity import ensure_local_identity
from app.services.pdf_chunks import persist_pdf_chunks
from app.services.pdfs import extract_pdf
from app.services.product_truth import ingest_product_truth_file
from app.services.retrieval import (
    _cosine_similarity,
    build_retrieval_request,
    build_retrieval_run,
)
from app.services.seed import import_historical_corpus
from app.services.storage import LocalObjectStorage
from app.services.workbooks import HISTORICAL_SCHEMA_VERSION, parse_workbook_bytes
from tests.seed_paths import historical_customer_dir, product_truth_path, seed_data_root


def sample_pdf_bytes(repo_root: Path) -> bytes:
    return (
        historical_customer_dir(repo_root, "nordtransit_logistik_ag")
        / "nordtransit_logistik_ag_context_brief.pdf"
    ).read_bytes()


def sample_xlsx_bytes(repo_root: Path) -> bytes:
    return (
        historical_customer_dir(repo_root, "nordtransit_logistik_ag")
        / "nordtransit_logistik_ag_qa.xlsx"
    ).read_bytes()


def create_invalid_header_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "QA"
    worksheet["A1"] = "Context"
    worksheet["B1"] = "Question Text"
    worksheet["C1"] = "Answer"
    worksheet["A2"] = "ctx"
    worksheet["B2"] = "q"
    worksheet["C2"] = "a"
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_case_with_questionnaire(
    session: Session,
    *,
    repo_root: Path,
    tenant_id,
    user_id,
    storage: LocalObjectStorage,
    ai_service: StubAIService,
    name: str = "NordTransit Pilot",
) -> RfxCase:
    return create_case_from_uploads(
        session,
        storage=storage,
        ai_service=ai_service,
        tenant_id=tenant_id,
        user_id=user_id,
        case_name=name,
        client_name="NordTransit Logistik AG",
        pdf_file_name="nordtransit_logistik_ag_context_brief.pdf",
        pdf_media_type="application/pdf",
        pdf_payload=sample_pdf_bytes(repo_root),
        questionnaire_file_name="nordtransit_logistik_ag_qa.xlsx",
        questionnaire_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        questionnaire_payload=sample_xlsx_bytes(repo_root),
    )


def create_manual_case_with_profile(
    session: Session,
    *,
    tenant_id,
    user_id,
    ai_service: StubAIService,
    case_name: str = "Manual Case",
    client_name: str = "Manual Client",
    language: str = "en",
    row_context: str = "Manual row context",
    row_question: str = "What is the requirement?",
    questionnaire_file_name: str = "questionnaire.xlsx",
    questionnaire_file_hash: str = "questionnaire-hash",
    case_pdf_file_name: str = "context.pdf",
    case_pdf_file_hash: str = "context-pdf-hash",
    case_fact_items: list[tuple[str, str]] | None = None,
    pdf_page_texts: list[str] | None = None,
) -> tuple[RfxCase, QuestionnaireRow, ChatThread]:
    pipeline = resolve_pipeline_selection(build_settings(env_file=None))
    artifact_hashes = artifact_index_hashes(pipeline)
    case = RfxCase(
        tenant_id=tenant_id,
        created_by_user_id=user_id,
        name=case_name,
        client_name=client_name,
        language=language,
        pipeline_profile_name=pipeline.profile_name,
        pipeline_config_json=pipeline.resolved_config,
        pipeline_config_hash=pipeline.config_hash,
        index_config_hash=pipeline.index_config_hash,
    )
    session.add(case)
    session.flush()
    pdf_upload = Upload(
        tenant_id=tenant_id,
        case_id=case.id,
        kind=UploadKind.CASE_PDF,
        original_file_name=case_pdf_file_name,
        media_type="application/pdf",
        object_key=f"tests/{case.id}/case-pdf",
        file_hash=case_pdf_file_hash,
        size_bytes=1,
        payload=b"x",
    )
    questionnaire_upload = Upload(
        tenant_id=tenant_id,
        case_id=case.id,
        kind=UploadKind.QUESTIONNAIRE_XLSX,
        original_file_name=questionnaire_file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        object_key=f"tests/{case.id}/questionnaire",
        file_hash=questionnaire_file_hash,
        size_bytes=1,
        payload=b"x",
    )
    session.add_all([pdf_upload, questionnaire_upload])
    session.flush()
    profile = CaseProfile(
        tenant_id=tenant_id,
        case_id=case.id,
        source_pdf_upload_id=pdf_upload.id,
        schema_version="rfx_case_profile.v3",
        prompt_set_version="rfx_case_profile_prompt_set.v3",
        model="stub-ai-service",
        summary=f"Case summary for {client_name}",
        source_file_name=case_pdf_file_name,
        source_file_hash=case_pdf_file_hash,
        language=language,
        pipeline_profile_name=pipeline.profile_name,
        index_config_json=case_profile_index_payload(pipeline.resolved_pipeline),
        index_config_hash=artifact_hashes.case_profile,
        generated_at=datetime.now(UTC),
        document={"case_id": str(case.id)},
    )
    session.add(profile)
    session.flush()
    for position, (analysis_item_id, answer) in enumerate(
        case_fact_items
        or [
            ("strategic_objectives", "Default strategic objective"),
            ("security_privacy_regulatory", "Default security requirement"),
        ],
        start=1,
    ):
        prompt = f"Prompt for {analysis_item_id}"
        session.add(
            CaseProfileItem(
                tenant_id=tenant_id,
                case_profile_id=profile.id,
                case_id=case.id,
                analysis_item_id=analysis_item_id,
                position=position,
                prompt=prompt,
                answer=answer,
                confidence="high",
                citations=["citation_unavailable"],
                normalized_text=f"{prompt}\n{answer}",
                embedding=ai_service.embed_text(f"{prompt}\n{answer}"),
            )
        )
    page_models: list[PdfPage] = []
    for index, text in enumerate(
        pdf_page_texts
        or ["Default PDF page about integrations and current-state process requirements."],
        start=1,
    ):
        page_models.append(
            PdfPage(
                tenant_id=tenant_id,
                case_id=case.id,
                upload_id=pdf_upload.id,
                page_number=index,
                extracted_text=text,
                text_hash=sha256_text(text),
            )
        )
    session.add_all(page_models)
    session.flush()
    persist_pdf_chunks(
        session,
        ai_service=ai_service,
        pipeline=pipeline,
        tenant_id=tenant_id,
        case_id=case.id,
        upload_id=pdf_upload.id,
        pages=page_models,
        case_name=case.name,
        client_name=case.client_name,
        language=case.language,
        source_file_name=pdf_upload.original_file_name,
    )
    questionnaire = Questionnaire(
        tenant_id=tenant_id,
        case_id=case.id,
        upload_id=questionnaire_upload.id,
        source_file_name=questionnaire_file_name,
        source_sheet_name="QA",
        file_hash=questionnaire_file_hash,
        schema_version="questionnaire_workbook.v1",
    )
    session.add(questionnaire)
    session.flush()
    row = QuestionnaireRow(
        tenant_id=tenant_id,
        case_id=case.id,
        questionnaire_id=questionnaire.id,
        source_sheet_name="QA",
        source_row_number=2,
        source_row_id=f"{questionnaire_file_name}:QA:2",
        context_raw=row_context,
        question_raw=row_question,
        answer_raw="",
        normalized_text=f"{row_context}\n{row_question}",
    )
    session.add(row)
    session.flush()
    thread = ChatThread(
        tenant_id=tenant_id,
        case_id=case.id,
        questionnaire_row_id=row.id,
        created_by_user_id=user_id,
        title="Thread",
    )
    session.add(thread)
    session.flush()
    return case, row, thread


def create_historical_package(
    session: Session,
    *,
    tenant_id,
    ai_service: StubAIService,
    dataset: HistoricalDataset,
    client_slug: str,
    client_name: str,
    language: str,
    workbook_name: str,
    workbook_hash: str,
    pdf_file_name: str,
    pdf_file_hash: str,
    case_summary: str,
    row_context: str,
    row_question: str,
    row_answer: str,
    source_row_id: str,
) -> HistoricalQARow:
    pipeline = resolve_pipeline_selection(build_settings(env_file=None))
    artifact_hashes = artifact_index_hashes(pipeline)
    workbook = HistoricalWorkbook(
        tenant_id=tenant_id,
        dataset_id=dataset.id,
        client_name=client_name,
        language=language,
        source_file_name=workbook_name,
        source_sheet_name="QA",
        file_hash=workbook_hash,
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(workbook)
    session.flush()
    package = HistoricalClientPackage(
        tenant_id=tenant_id,
        dataset_id=dataset.id,
        workbook_id=workbook.id,
        client_slug=client_slug,
        client_name=client_name,
        language=language,
        source_pdf_file_name=pdf_file_name,
        source_pdf_file_hash=pdf_file_hash,
        pdf_object_key=f"tests/historical/{client_slug}.pdf",
        pdf_size_bytes=1,
        pipeline_profile_name=pipeline.profile_name,
        index_config_json=historical_index_payload(pipeline.resolved_pipeline),
        index_config_hash=artifact_hashes.historical,
    )
    session.add(package)
    session.flush()
    profile = HistoricalCaseProfile(
        tenant_id=tenant_id,
        client_package_id=package.id,
        schema_version="rfx_case_profile.v3",
        prompt_set_version="rfx_case_profile_prompt_set.v3",
        model="stub-ai-service",
        summary=case_summary,
        source_file_name=pdf_file_name,
        source_file_hash=pdf_file_hash,
        language=language,
        generated_at=datetime.now(UTC),
        signature_version="case_signature.v1",
        signature_embedding_model="stub-ai-service",
        signature_fields_json={
            "summary": case_summary,
            "analysis_item_ids": ["strategic_objectives"],
        },
        signature_text=case_summary,
        signature_embedding=ai_service.embed_text(case_summary),
        document={"case_id": str(package.id)},
    )
    session.add(profile)
    session.flush()
    session.add(
        HistoricalCaseProfileItem(
            tenant_id=tenant_id,
            historical_case_profile_id=profile.id,
            analysis_item_id="strategic_objectives",
            position=1,
            prompt="Prompt",
            answer=case_summary,
            confidence="high",
            citations=["citation_unavailable"],
            normalized_text=case_summary,
            embedding=ai_service.embed_text(case_summary),
        )
    )
    row = HistoricalQARow(
        tenant_id=tenant_id,
        dataset_id=dataset.id,
        workbook_id=workbook.id,
        client_package_id=package.id,
        historical_case_profile_id=profile.id,
        client_name=client_name,
        source_file_name=workbook_name,
        source_sheet_name="QA",
        source_row_number=2,
        source_row_id=source_row_id,
        language=language,
        language_confidence=1.0,
        approval_status=ApprovalStatus.APPROVED,
        context_raw=row_context,
        question_raw=row_question,
        answer_raw=row_answer,
        normalized_text=f"{row_context}\n{row_question}",
        file_hash=workbook_hash,
        schema_version=HISTORICAL_SCHEMA_VERSION,
        embedding=ai_service.embed_text(f"{row_context}\n{row_question}"),
    )
    session.add(row)
    session.flush()
    return row


class AmbiguousTruthVector:
    def __init__(self, values: list[float]) -> None:
        self._values = values

    def __iter__(self):
        return iter(self._values)

    def __bool__(self) -> bool:
        raise ValueError("Truthiness must not be used for vector-like objects.")


def test_malformed_workbook_schema_fails_loudly() -> None:
    with pytest.raises(ValidationFailure, match="exact headers"):
        parse_workbook_bytes(
            create_invalid_header_workbook(),
            source_file_name="bad.xlsx",
            schema_version=HISTORICAL_SCHEMA_VERSION,
            allow_empty_answer=False,
        )


def test_bluepeak_seed_import_uses_exact_sample_files(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    dataset = import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
    )
    session.commit()

    workbook_names = {
        workbook.source_file_name
        for workbook in session.scalars(
            select(HistoricalWorkbook).where(HistoricalWorkbook.dataset_id == dataset.id)
        ).all()
    }
    assert workbook_names == {
        "asteron_industrial_components_bv_qa.xlsx",
        "crownshield_insurance_services_ltd_qa.xlsx",
        "nordtransit_logistik_ag_qa.xlsx",
    }
    row_count = session.scalar(
        select(func.count())
        .select_from(HistoricalQARow)
        .where(HistoricalQARow.dataset_id == dataset.id)
    )
    assert row_count == 90
    first_row = session.scalar(
        select(HistoricalQARow)
        .where(HistoricalQARow.dataset_id == dataset.id)
        .order_by(HistoricalQARow.source_row_number.asc(), HistoricalQARow.source_file_name.asc())
    )
    assert first_row is not None
    assert first_row.source_row_id.endswith(":QA:2")
    assert first_row.file_hash
    assert first_row.client_package_id is not None
    assert first_row.historical_case_profile_id is not None
    package_count = session.scalar(
        select(func.count()).select_from(HistoricalClientPackage).where(
            HistoricalClientPackage.dataset_id == dataset.id
        )
    )
    profile_count = session.scalar(
        select(func.count()).select_from(HistoricalCaseProfile).where(
            HistoricalCaseProfile.tenant_id == context.tenant.id
        )
    )
    assert package_count == 3
    assert profile_count == 3
    first_profile = session.get(HistoricalCaseProfile, first_row.historical_case_profile_id)
    assert first_profile is not None
    assert first_profile.signature_version == "case_signature.v1"


def test_bluepeak_seed_reimport_replaces_existing_dataset_without_upload_fk_breakage(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    session.execute(sql_text("PRAGMA foreign_keys=ON"))
    context = ensure_local_identity(session, settings)
    first = import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
    )
    session.flush()
    second = import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
    )
    session.flush()
    dataset_count = session.scalar(
        select(func.count()).select_from(HistoricalDataset).where(
            HistoricalDataset.tenant_id == context.tenant.id,
            HistoricalDataset.slug == "sample-historical-corpus",
        )
    )
    assert first.slug == second.slug == "sample-historical-corpus"
    assert dataset_count == 1


def test_invalid_case_profile_schema_fails_loudly(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    class InvalidCaseProfileAIService(StubAIService):
        def generate_case_profile(self, **kwargs):  # type: ignore[override]
            document = super().generate_case_profile(**kwargs)
            return CaseProfileDocument.model_construct(
                **{
                    **document.model_dump(mode="python"),
                    "source_pdf": document.source_pdf,
                    "analysis_items": document.analysis_items[:-1],
                }
            )

    context = ensure_local_identity(session, settings)
    case = RfxCase(
        tenant_id=context.tenant.id,
        created_by_user_id=context.user.id,
        name="Broken Profile Case",
        client_name="NordTransit Logistik AG",
        language="de",
    )
    session.add(case)
    session.flush()
    pdf_payload = sample_pdf_bytes(repo_root)
    extracted = extract_pdf(pdf_payload)
    upload = Upload(
        tenant_id=context.tenant.id,
        case_id=case.id,
        kind=UploadKind.CASE_PDF,
        original_file_name="nordtransit_logistik_ag_context_brief.pdf",
        media_type="application/pdf",
        object_key="tests/broken.pdf",
        file_hash=extracted.file_hash,
        size_bytes=len(pdf_payload),
        payload=pdf_payload,
    )
    session.add(upload)
    session.flush()
    from app.models.entities import PdfPage

    page_models = [
        PdfPage(
            tenant_id=context.tenant.id,
            case_id=case.id,
            upload_id=upload.id,
            page_number=page.page_number,
            extracted_text=page.text,
            text_hash=page.text_hash,
        )
        for page in extracted.pages
    ]
    session.add_all(page_models)
    session.flush()

    with pytest.raises(ValidationFailure, match="case_profile failed schema validation"):
        persist_case_profile(
            session,
            ai_service=InvalidCaseProfileAIService(),
            pipeline=resolve_pipeline_selection(settings),
            case=case,
            upload=upload,
            pdf_pages=page_models,
        )


def test_retrieval_cannot_cross_tenant_boundaries(session: Session, settings: Settings) -> None:
    local_context = ensure_local_identity(session, settings)
    other_tenant = Tenant(name="Other Tenant", slug="other-tenant")
    other_user = User(email="other.user@example.com", display_name="Other User")
    session.add_all([other_tenant, other_user])
    session.flush()
    from app.models.entities import Membership

    session.add(
        Membership(
            tenant_id=other_tenant.id,
            user_id=other_user.id,
            role=MembershipRole.ADMIN,
        )
    )
    session.flush()

    sample_dataset = HistoricalDataset(
        tenant_id=local_context.tenant.id,
        name="Sample",
        slug="sample",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    other_dataset = HistoricalDataset(
        tenant_id=other_tenant.id,
        name="Other",
        slug="other",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add_all([sample_dataset, other_dataset])
    session.flush()

    ai_service = StubAIService()
    create_historical_package(
        session,
        tenant_id=local_context.tenant.id,
        ai_service=ai_service,
        dataset=sample_dataset,
        client_slug="sample-client",
        client_name="Sample Client",
        language="en",
        workbook_name="sample.xlsx",
        workbook_hash="hash-sample",
        pdf_file_name="sample.pdf",
        pdf_file_hash="pdf-hash-sample",
        case_summary="Logistics workflow modernization and audit controls.",
        row_context="Logistics operations and compliance context",
        row_question="What evidence exists?",
        row_answer="Tenant one evidence",
        source_row_id="sample.xlsx:QA:2",
    )
    create_historical_package(
        session,
        tenant_id=other_tenant.id,
        ai_service=ai_service,
        dataset=other_dataset,
        client_slug="other-client",
        client_name="Other Client",
        language="en",
        workbook_name="other.xlsx",
        workbook_hash="hash-other",
        pdf_file_name="other.pdf",
        pdf_file_hash="pdf-hash-other",
        case_summary="Insurance claims workflow and delegated authority controls.",
        row_context="Insurance operations and compliance context",
        row_question="What evidence exists?",
        row_answer="Tenant two evidence",
        source_row_id="other.xlsx:QA:2",
    )
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=local_context.tenant.id,
        user_id=local_context.user.id,
        ai_service=ai_service,
        case_name="Tenant One Case",
        client_name="Sample Client",
        row_context="Logistics operations and compliance context",
        row_question="What evidence exists?",
        case_fact_items=[("strategic_objectives", "Logistics workflow modernization")],
        pdf_page_texts=["Logistics compliance evidence and rollout governance."],
    )

    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    _, evidence = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=local_context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )

    assert evidence
    assert all("Tenant two evidence" not in item.excerpt for item in evidence)


def test_live_row_context_changes_retrieval_ranking(session: Session, settings: Settings) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="GDPR, German hosting, and data residency requirements",
        row_question="How will the requirement be addressed?",
        case_fact_items=[
            ("security_privacy_regulatory", "GDPR controls and German hosting support."),
            ("architecture_integration_data", "SAP integration and API connectivity."),
        ],
        pdf_page_texts=[
            "Detailed GDPR and German hosting evidence.",
            "Detailed SAP integration evidence.",
        ],
    )
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    _, evidence = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    current_case_facts = [item for item in evidence if item.source_label == "current_case_facts"]
    assert current_case_facts[0].source_title == "security_privacy_regulatory"

    row.context_raw = "SAP integration landscape and ERP connectivity requirements"
    row.normalized_text = f"{row.context_raw}\n{row.question_raw}"
    session.flush()
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    _, evidence = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    current_case_facts = [item for item in evidence if item.source_label == "current_case_facts"]
    assert current_case_facts[0].source_title == "architecture_integration_data"


def test_historical_client_context_similarity_changes_exemplar_ranking(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        client_name="NordTransit Logistik AG",
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        case_fact_items=[("strategic_objectives", "Logistics workflow modernization for carriers.")],
    )
    dataset = HistoricalDataset(
        tenant_id=context.tenant.id,
        name="Sample",
        slug="sample-ranked",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(dataset)
    session.flush()
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="logistics",
        client_name="Logistics Client",
        language="en",
        workbook_name="logistics.xlsx",
        workbook_hash="hash-logistics",
        pdf_file_name="logistics.pdf",
        pdf_file_hash="pdf-logistics",
        case_summary="Logistics workflow modernization for carriers.",
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        row_answer="Logistics exemplar answer",
        source_row_id="logistics.xlsx:QA:2",
    )
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="insurance",
        client_name="Insurance Client",
        language="en",
        workbook_name="insurance.xlsx",
        workbook_hash="hash-insurance",
        pdf_file_name="insurance.pdf",
        pdf_file_hash="pdf-insurance",
        case_summary="Insurance claims adjudication and underwriting workflows.",
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        row_answer="Insurance exemplar answer",
        source_row_id="insurance.xlsx:QA:2",
    )
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    _, evidence = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    historical = [item for item in evidence if item.source_label == "historical_exemplar"]
    assert historical[0].metadata_json["provenance"]["client_slug"] == "logistics"


def test_historical_answer_text_is_not_primary_recall_field(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        case_fact_items=[("strategic_objectives", "Logistics workflow modernization for carriers.")],
    )
    dataset = HistoricalDataset(
        tenant_id=context.tenant.id,
        name="Sample",
        slug="sample-answer-text",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(dataset)
    session.flush()
    relevant_row = create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="logistics",
        client_name="Logistics Client",
        language="en",
        workbook_name="logistics.xlsx",
        workbook_hash="hash-logistics-answer",
        pdf_file_name="logistics.pdf",
        pdf_file_hash="pdf-logistics-answer",
        case_summary="Logistics workflow modernization for carriers.",
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        row_answer="Original logistics answer",
        source_row_id="logistics.xlsx:QA:2",
    )
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="insurance",
        client_name="Insurance Client",
        language="en",
        workbook_name="insurance.xlsx",
        workbook_hash="hash-insurance-answer",
        pdf_file_name="insurance.pdf",
        pdf_file_hash="pdf-insurance-answer",
        case_summary="Insurance claims adjudication and underwriting workflows.",
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        row_answer="Insurance answer",
        source_row_id="insurance.xlsx:QA:2",
    )
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    _, evidence_before = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    top_before = [item for item in evidence_before if item.source_label == "historical_exemplar"][0]

    relevant_row.answer_raw = "Completely different answer text that should not drive recall."
    session.flush()
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    _, evidence_after = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    top_after = [item for item in evidence_after if item.source_label == "historical_exemplar"][0]
    assert top_before.source_id == top_after.source_id == relevant_row.id


def test_same_source_leakage_is_excluded(session: Session, container, repo_root: Path, settings: Settings) -> None:
    context = ensure_local_identity(session, settings)
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
    )
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    profile = session.scalar(select(CaseProfile).where(CaseProfile.case_id == case.id))
    assert profile is not None
    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    historical_items = [item for item in result.evidence_items if item.source_label == "historical_exemplar"]
    assert historical_items
    for item in historical_items:
        provenance = item.metadata_json["provenance"]
        assert provenance["source_row_id"] != row.source_row_id
        assert provenance["source_file_name"] != "nordtransit_logistik_ag_qa.xlsx"
        assert provenance["source_pdf_file_hash"] != profile.source_file_hash


def test_style_only_revision_reuses_prior_retrieval_run_and_source_plan(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
    )
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    first = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    second = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Make it shorter",
        thread=first.thread,
    )
    first_run = session.get(ExecutionRun, first.answer_version.execution_run_id)
    second_run = session.get(ExecutionRun, second.answer_version.execution_run_id)
    assert first_run is not None and second_run is not None
    assert second.answer_version.retrieval_run_id == first.answer_version.retrieval_run_id
    assert second_run.outputs_json["generation_path"] == "render_only_reuse_plan"
    assert second_run.outputs_json["source_planning_model_invocation_id"] == first_run.outputs_json["planning_model_invocation_id"]
    assert second_run.outputs_json["reused_answer_version_id"] == str(first.answer_version.id)
    planning_invocations = [
        invocation
        for invocation in session.scalars(
            select(ModelInvocation).where(ModelInvocation.execution_run_id == second_run.id)
        ).all()
        if invocation.kind == ModelInvocationKind.ANSWER_GENERATION
        and invocation.metadata_json.get("prompt_family") == "answer_planning"
    ]
    assert planning_invocations == []
    assert [item.source_id for item in second.evidence_items] == [item.source_id for item in first.evidence_items]


def test_explicit_style_only_override_uses_render_only_revise_without_heuristic_rejection(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
    )
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    first = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    second = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Mention GDPR and German hosting.",
        thread=first.thread,
        revision_mode_override="style_only",
    )
    first_run = session.get(ExecutionRun, first.answer_version.execution_run_id)
    second_run = session.get(ExecutionRun, second.answer_version.execution_run_id)
    assert first_run is not None and second_run is not None
    assert second.answer_version.retrieval_run_id == first.answer_version.retrieval_run_id
    assert second_run.outputs_json["generation_path"] == "render_only_reuse_plan"
    assert second_run.outputs_json["source_planning_model_invocation_id"] == first_run.outputs_json["planning_model_invocation_id"]


def test_content_changing_revision_refreshes_retrieval(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
    )
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    first = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    second = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Mention GDPR and German hosting.",
        thread=first.thread,
    )
    second_run = session.get(RetrievalRun, second.answer_version.retrieval_run_id)
    assert second_run is not None
    assert second_run.request_context["retrieval_action"] == "refresh_retrieval"
    assert second_run.request_context["retrieval_action_reason"] == "new_or_content_change_requires_refresh"
    assert second_run.request_context["revision_mode"] == "content_change"
    assert second_run.request_context["revision_classifier"]["version"] == "revision_classifier.v2"
    assert second_run.request_context["revision_classifier"]["reason"] == "matched_content_change_pattern"
    assert second_run.request_context["reused_from_retrieval_run_id"] is None
    assert second_run.request_context["feature_texts"]["revision_intent"] == "Mention GDPR and German hosting."


def test_raw_current_pdf_evidence_is_retrieved_as_distinct_layer(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Carrier onboarding portal and multilingual supplier forms",
        row_question="How will the portal work?",
        pdf_page_texts=[
            "Carrier onboarding portal with multilingual supplier forms and document intake.",
            "Unrelated analytics page.",
        ],
    )
    persisted_chunks = session.scalars(
        select(PdfChunk)
        .where(PdfChunk.case_id == case.id)
        .order_by(PdfChunk.page_number.asc(), PdfChunk.chunk_index.asc())
    ).all()
    assert persisted_chunks
    assert persisted_chunks[0].chunking_version == "pdf_chunker.v1"
    assert persisted_chunks[0].embedding_model == settings.openai_embedding_model
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    _, evidence = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    raw_pdf_items = [item for item in evidence if item.source_label == "raw_current_pdf"]
    assert raw_pdf_items
    assert raw_pdf_items[0].source_kind.value == "pdf_chunk"
    assert raw_pdf_items[0].source_title == "page 1 chunk 1"
    assert raw_pdf_items[0].metadata_json["provenance"]["chunking_version"] == "pdf_chunker.v1"


def test_candidate_generation_broadening_is_explicit(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="zzzz unmatched keyword set",
        row_question="qqqq unmatched question",
        pdf_page_texts=["General process requirements without matching keywords."],
    )
    dataset = HistoricalDataset(
        tenant_id=context.tenant.id,
        name="Sample",
        slug="sample-broaden",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(dataset)
    session.flush()
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="historic",
        client_name="Historic Client",
        language="en",
        workbook_name="historic.xlsx",
        workbook_hash="historic-hash",
        pdf_file_name="historic.pdf",
        pdf_file_hash="historic-pdf-hash",
        case_summary="Historic operations summary",
        row_context="Historic context",
        row_question="Historic question",
        row_answer="Historic answer",
        source_row_id="historic.xlsx:QA:2",
    )
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    retrieval_run, _ = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    execution = retrieval_run.request_context["retrieval_execution"]
    assert execution["broadened"] is True
    broadened_stages = [
        stage for stage in execution["stages"] if stage["stage"] == "broadened_scope"
    ]
    assert broadened_stages
    assert any(stage["reason"] == "keyword/vector stages produced zero candidates" for stage in broadened_stages)


def test_retrieval_sufficiency_is_explicit_when_supporting_corpora_are_missing(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Carrier onboarding portal and multilingual supplier forms",
        row_question="How will the portal work?",
        pdf_page_texts=[
            "Carrier onboarding portal with multilingual supplier forms and document intake.",
        ],
    )
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    retrieval_run, _ = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    sufficiency = retrieval_run.request_context["retrieval_execution"]["sufficiency"]
    assert sufficiency["status"] == "degraded"
    assert sufficiency["degraded"] is True
    assert "No product-truth evidence was selected." in sufficiency["notes"]


def test_historical_case_signature_provenance_mismatch_fails_loudly(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
    )
    dataset = HistoricalDataset(
        tenant_id=context.tenant.id,
        name="Sample",
        slug="sample-signature-mismatch",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(dataset)
    session.flush()
    first_row = create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="first",
        client_name="First Client",
        language="en",
        workbook_name="first.xlsx",
        workbook_hash="first-hash",
        pdf_file_name="first.pdf",
        pdf_file_hash="first-pdf-hash",
        case_summary="First summary",
        row_context="First context",
        row_question="First question",
        row_answer="First answer",
        source_row_id="first.xlsx:QA:2",
    )
    second_row = create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="second",
        client_name="Second Client",
        language="en",
        workbook_name="second.xlsx",
        workbook_hash="second-hash",
        pdf_file_name="second.pdf",
        pdf_file_hash="second-pdf-hash",
        case_summary="Second summary",
        row_context="Second context",
        row_question="Second question",
        row_answer="Second answer",
        source_row_id="second.xlsx:QA:2",
    )
    first_row.historical_case_profile_id = second_row.historical_case_profile_id
    session.flush()
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    with pytest.raises(ValidationFailure, match="points to a historical case profile from another package"):
        build_retrieval_run(
            session,
            ai_service=ai_service,
            tenant_id=context.tenant.id,
            case_id=case.id,
            row=row,
            thread_id=thread.id,
            request=request,
        )


def test_provenance_based_dedup_is_explicit(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        case_fact_items=[("strategic_objectives", "Logistics workflow modernization for carriers.")],
    )
    dataset = HistoricalDataset(
        tenant_id=context.tenant.id,
        name="Sample",
        slug="sample-dedup",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(dataset)
    session.flush()
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="dup-a",
        client_name="Duplicate A",
        language="en",
        workbook_name="dup-a.xlsx",
        workbook_hash="dup-a-hash",
        pdf_file_name="shared.pdf",
        pdf_file_hash="shared-pdf-hash",
        case_summary="Logistics workflow modernization for carriers.",
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        row_answer="Duplicate answer A",
        source_row_id="shared.xlsx:QA:2",
    )
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="dup-b",
        client_name="Duplicate B",
        language="en",
        workbook_name="dup-b.xlsx",
        workbook_hash="dup-b-hash",
        pdf_file_name="shared.pdf",
        pdf_file_hash="shared-pdf-hash",
        case_summary="Logistics workflow modernization for carriers.",
        row_context="Carrier collaboration workflow and logistics exception handling",
        row_question="How would you support the workflow?",
        row_answer="Duplicate answer B",
        source_row_id="shared.xlsx:QA:2",
    )
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    retrieval_run, evidence = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    historical = [item for item in evidence if item.source_label == "historical_exemplar"]
    assert len(historical) == 1
    dedup = retrieval_run.request_context["retrieval_execution"]["dedup"]
    assert dedup
    assert dedup[0]["source_label"] == "historical_exemplar"


def test_same_language_exemplar_preference_is_explicit(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        language="de",
        row_context="Datenschutz und Lieferantenportal",
        row_question="Wie wird der Prozess unterstützt?",
        case_fact_items=[("security_privacy_regulatory", "Datenschutz und deutsches Hosting.")],
    )
    dataset = HistoricalDataset(
        tenant_id=context.tenant.id,
        name="Sample",
        slug="sample-language",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(dataset)
    session.flush()
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="de-client",
        client_name="German Client",
        language="de",
        workbook_name="de.xlsx",
        workbook_hash="hash-de",
        pdf_file_name="de.pdf",
        pdf_file_hash="pdf-de",
        case_summary="Datenschutz und deutsches Lieferantenportal.",
        row_context="Datenschutz und Lieferantenportal",
        row_question="Wie wird der Prozess unterstützt?",
        row_answer="Deutsches Beispiel",
        source_row_id="de.xlsx:QA:2",
    )
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        dataset=dataset,
        client_slug="en-client",
        client_name="English Client",
        language="en",
        workbook_name="en.xlsx",
        workbook_hash="hash-en",
        pdf_file_name="en.pdf",
        pdf_file_hash="pdf-en",
        case_summary="Data privacy and supplier portal.",
        row_context="Datenschutz und Lieferantenportal",
        row_question="Wie wird der Prozess unterstützt?",
        row_answer="English example",
        source_row_id="en.xlsx:QA:2",
    )
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Entwurf erstellen.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    _, evidence = build_retrieval_run(
        session,
        ai_service=ai_service,
        tenant_id=context.tenant.id,
        case_id=case.id,
        row=row,
        thread_id=thread.id,
        request=request,
    )
    historical = [item for item in evidence if item.source_label == "historical_exemplar"]
    assert historical[0].metadata_json["language"] == "de"
    assert historical[0].metadata_json["cross_lingual_fallback"] is False


def test_missing_historical_case_profile_provenance_fails_loudly(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row, thread = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Compliance workflow requirements",
        row_question="How would you support the workflow?",
    )
    dataset = HistoricalDataset(
        tenant_id=context.tenant.id,
        name="Broken Sample",
        slug="broken-sample",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(dataset)
    session.flush()
    workbook = HistoricalWorkbook(
        tenant_id=context.tenant.id,
        dataset_id=dataset.id,
        client_name="Broken Client",
        language="en",
        source_file_name="broken.xlsx",
        source_sheet_name="QA",
        file_hash="broken-hash",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(workbook)
    session.flush()
    session.add(
        HistoricalQARow(
            tenant_id=context.tenant.id,
            dataset_id=dataset.id,
            workbook_id=workbook.id,
            client_package_id=None,
            client_name="Broken Client",
            source_file_name="broken.xlsx",
            source_sheet_name="QA",
            source_row_number=2,
            source_row_id="broken.xlsx:QA:2",
            language="en",
            language_confidence=1.0,
            approval_status=ApprovalStatus.APPROVED,
            context_raw="Compliance workflow requirements",
            question_raw="How would you support the workflow?",
            answer_raw="Broken answer",
            normalized_text="Compliance workflow requirements\nHow would you support the workflow?",
            file_hash="broken-hash",
            schema_version=HISTORICAL_SCHEMA_VERSION,
            embedding=ai_service.embed_text(
                "Compliance workflow requirements\nHow would you support the workflow?"
            ),
        )
    )
    session.flush()
    request = build_retrieval_request(
        session,
        case=case,
        row=row,
        user_message="Draft the answer.",
        revision_mode="initial_draft",
        retrieval_action="refresh_retrieval",
        previous_answer_text=None,
    )
    with pytest.raises(ValidationFailure, match="missing historical client-package provenance"):
        build_retrieval_run(
            session,
            ai_service=ai_service,
            tenant_id=context.tenant.id,
            case_id=case.id,
            row=row,
            thread_id=thread.id,
            request=request,
        )


def test_conversations_cannot_mix_cases(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    first_case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
        name="Case One",
    )
    second_case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
        name="Case Two",
    )
    first_row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == first_case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    second_row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == second_case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert first_row is not None and second_row is not None
    foreign_thread = ChatThread(
        tenant_id=context.tenant.id,
        case_id=second_case.id,
        questionnaire_row_id=second_row.id,
        created_by_user_id=context.user.id,
        title="Foreign thread",
    )
    session.add(foreign_thread)
    session.flush()

    with pytest.raises(ScopeViolation, match="cannot mix cases"):
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=first_case,
            row=first_row,
            user_id=context.user.id,
            user_message="Revise this answer.",
            thread=foreign_thread,
        )


def test_exported_rows_map_deterministically(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    rows = session.scalars(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    ).all()
    assert rows
    for row in rows:
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the row answer.",
            thread=None,
        )
    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    assert questionnaire is not None
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None
    from app.services.exports import export_questionnaire

    export_job = export_questionnaire(
        session,
        storage=container.storage,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
    )
    assert rows[0].source_row_id in export_job.row_mapping_json
    download_upload = session.get(Upload, export_job.output_upload_id)
    assert download_upload is not None
    workbook = load_workbook(BytesIO(container.storage.read_bytes(download_upload.object_key)))
    worksheet = workbook[questionnaire.source_sheet_name]
    assert (
        worksheet.cell(row=rows[0].source_row_number, column=3).value
        == session.scalar(
            select(AnswerVersion.answer_text)
            .where(AnswerVersion.questionnaire_row_id == rows[0].id)
            .order_by(AnswerVersion.version_number.desc())
        )
    )


def test_answer_versions_capture_raw_prompt_and_response(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Explain which products match the scope.",
        thread=None,
    )
    assert result.answer_version.llm_capture_status == LLMCaptureStatus.CAPTURED
    assert result.answer_version.llm_request_text is not None
    assert "<question>" in result.answer_version.llm_request_text
    assert "<answer_plan_json>" in result.answer_version.llm_request_text
    assert result.answer_version.llm_response_text == result.answer_version.answer_text


def test_style_only_revision_render_prompt_replays_prior_thread_messages(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    first = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    second = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Make it shorter.",
        thread=first.thread,
    )
    third = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Remove the bullet points.",
        thread=first.thread,
        revision_mode_override="style_only",
    )
    prompt = third.answer_version.llm_request_text
    assert prompt is not None
    assert "Draft the answer." in prompt
    assert json.dumps(first.answer_version.answer_text, ensure_ascii=True)[1:-1] in prompt
    assert "Make it shorter." in prompt
    assert json.dumps(second.answer_version.answer_text, ensure_ascii=True)[1:-1] in prompt
    assert "<user_request>" in prompt
    assert "Remove the bullet points." in prompt
    assert '"role":"assistant"' in prompt
    assert '"type":"output_text"' in prompt


def test_thread_detail_orders_same_timestamp_messages_as_user_then_assistant(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
    client,
    auth_headers: dict[str, str],
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.desc())
    )
    assert row is not None
    thread = ChatThread(
        tenant_id=case.tenant_id,
        case_id=case.id,
        questionnaire_row_id=row.id,
        created_by_user_id=context.user.id,
        title="Ordering regression thread",
    )
    session.add(thread)
    session.flush()
    turns = [
        (
            datetime(2026, 3, 13, 12, 20, 54, tzinfo=UTC),
            "Generate a grounded answer for this row.",
            "We support structured operational processes with a repeatable, governed delivery approach.",
        ),
        (
            datetime(2026, 3, 13, 12, 21, 49, tzinfo=UTC),
            "make this a heiku",
            "We support structured operational processes with a repeatable, governed delivery approach in verse.",
        ),
        (
            datetime(2026, 3, 13, 12, 24, 7, tzinfo=UTC),
            "please make the answer a japanese heiku",
            "Configured first, then grow; governed standards, local variants; pilots lead the way.",
        ),
    ]
    for stamp, user_content, assistant_content in turns:
        # Insert assistant before user to reproduce timestamp-tie instability.
        session.add(
            ChatMessage(
                tenant_id=case.tenant_id,
                case_id=case.id,
                questionnaire_row_id=row.id,
                thread_id=thread.id,
                role=MessageRole.ASSISTANT,
                content=assistant_content,
                created_at=stamp,
            )
        )
        session.add(
            ChatMessage(
                tenant_id=case.tenant_id,
                case_id=case.id,
                questionnaire_row_id=row.id,
                thread_id=thread.id,
                role=MessageRole.USER,
                content=user_content,
                created_at=stamp,
            )
        )
    session.commit()

    response = client.get(
        f"/api/cases/{case.id}/threads/{thread.id}",
        headers=auth_headers,
    )
    assert response.status_code == 200
    body = response.json()
    assert [message["role"] for message in body["messages"]] == [
        "user",
        "assistant",
        "user",
        "assistant",
        "user",
        "assistant",
    ]
    assert body["messages"][-1]["content"].startswith("Configured first, then grow;")


def test_content_change_revision_render_prompt_does_not_replay_prior_thread_messages(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    first = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    second = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Mention GDPR and German hosting.",
        thread=first.thread,
    )
    prompt = second.answer_version.llm_request_text
    assert prompt is not None
    assert "Draft the answer." not in prompt
    assert json.dumps(first.answer_version.answer_text, ensure_ascii=True)[1:-1] not in prompt
    assert "Mention GDPR and German hosting." in prompt


def test_repeated_exports_produce_distinct_artifacts(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    rows = session.scalars(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    ).all()
    assert rows
    for row in rows:
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the row answer.",
            thread=None,
        )
    questionnaire = session.scalar(
        select(Questionnaire).where(Questionnaire.case_id == case.id)
    )
    assert questionnaire is not None
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None
    from app.services.exports import export_questionnaire

    first_export = export_questionnaire(
        session,
        storage=container.storage,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
    )
    second_export = export_questionnaire(
        session,
        storage=container.storage,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
    )
    assert first_export.output_upload_id != second_export.output_upload_id


def test_missing_export_provenance_fails_loudly(
    client,
    session_factory,
    container,
    auth_headers: dict[str, str],
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_questionnaire(
            session,
            repo_root=repo_root,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            storage=container.storage,
            ai_service=StubAIService(),
        )
        questionnaire = session.scalar(
            select(Questionnaire).where(Questionnaire.case_id == case.id)
        )
        assert questionnaire is not None
        session.execute(delete(Upload).where(Upload.id == questionnaire.upload_id))
        session.commit()

    response = client.post(
        f"/api/cases/{case.id}/export",
        headers=auth_headers,
        json={"mode": "latest_available"},
    )
    assert response.status_code == 422
    assert "missing its source upload provenance" in response.json()["detail"]


def test_dev_table_browser_rejects_unknown_tables(
    client,
    auth_headers: dict[str, str],
) -> None:
    tables = client.get("/api/dev/tables", headers=auth_headers)
    assert tables.status_code == 200
    names = {item["name"] for item in tables.json()["tables"]}
    assert {"execution_runs", "model_invocations", "artifact_builds"} <= names
    response = client.get("/api/dev/tables/not_a_real_table", headers=auth_headers)
    assert response.status_code == 422
    assert "Unsupported dev table" in response.json()["detail"]


def test_dev_table_browser_exposes_prompt_capture_fields(
    client,
    session_factory,
    container,
    auth_headers: dict[str, str],
    repo_root: Path,
    settings: Settings,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_questionnaire(
            session,
            repo_root=repo_root,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            storage=container.storage,
            ai_service=StubAIService(),
        )
        row = session.scalar(
            select(QuestionnaireRow)
            .where(QuestionnaireRow.case_id == case.id)
            .order_by(QuestionnaireRow.source_row_number.asc())
        )
        assert row is not None
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Create an answer draft.",
            thread=None,
        )
        session.commit()
        case_id = case.id

    response = client.get(
        f"/api/dev/tables/answer_versions?case_id={case_id}",
        headers=auth_headers,
    )
    assert response.status_code == 200
    row = response.json()["rows"][0]
    assert row["llm_capture_status"] == "captured"
    assert "<question>" in row["llm_request_text"]
    assert row["llm_response_text"] == row["answer_text"]


def test_dev_table_browser_rejects_case_filter_for_tenant_scoped_tables(
    client,
    session_factory,
    auth_headers: dict[str, str],
    repo_root: Path,
    settings: Settings,
    container,
) -> None:
    with session_factory() as session:
        context = ensure_local_identity(session, settings)
        case = create_case_with_questionnaire(
            session,
            repo_root=repo_root,
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            storage=container.storage,
            ai_service=StubAIService(),
        )
        session.commit()
        case_id = case.id

    response = client.get(
        f"/api/dev/tables/product_truth_records?case_id={case_id}",
        headers=auth_headers,
    )
    assert response.status_code == 422
    assert "does not support case filtering" in response.json()["detail"]


def test_unknown_vendor_claims_remain_explicit_unknowns(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="State supported deployment models.",
        thread=None,
    )
    assert (
        "Offen bleibt" in result.answer_version.answer_text
        or "An open point is" in result.answer_version.answer_text
    )


def test_live_case_profile_version_mismatch_fails_loudly(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    profile = session.scalar(
        select(CaseProfile)
        .where(CaseProfile.case_id == case.id)
        .order_by(CaseProfile.created_at.desc())
    )
    assert row is not None
    assert profile is not None
    profile.prompt_set_version = "rfx_case_profile_prompt_set.v1"
    session.flush()
    with pytest.raises(ValidationFailure, match="prompt set version"):
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
            reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
        )


def test_sample_product_truth_source_is_ingested_and_retrieved_separately(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    records = ingest_product_truth_file(
        session,
        ai_service=StubAIService(),
        tenant_id=context.tenant.id,
        path=product_truth_path(repo_root),
    )
    assert len(records) == 4
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
    )
    case = create_case_with_questionnaire(
        session,
        repo_root=repo_root,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        storage=container.storage,
        ai_service=StubAIService(),
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Which BluePeak products match the scope?",
        thread=None,
    )
    labels = {item.source_label for item in result.evidence_items}
    assert "current_case_facts" in labels
    assert "product_truth" in labels
    assert "historical_exemplar" in labels
    assert "remain unknown" not in result.answer_version.answer_text


def test_cosine_similarity_accepts_vector_like_values_without_truthiness() -> None:
    score = _cosine_similarity(
        [1.0, 0.0, 0.0],
        AmbiguousTruthVector([1.0, 0.0, 0.0]),
    )
    assert score == pytest.approx(1.0)
