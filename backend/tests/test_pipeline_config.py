from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.config import build_settings
from app.exceptions import ValidationFailure
from app.models.entities import (
    ArtifactBuild,
    BulkFillRequest,
    CaseProfile,
    ExecutionRun,
    HistoricalClientPackage,
    HistoricalQARow,
    ModelInvocation,
    PdfChunk,
    ProductTruthRecord,
    Questionnaire,
    QuestionnaireRow,
    RetrievalRun,
    RfxCase,
)
from app.models.enums import ArtifactBuildStatus, ModelInvocationKind
from app.pipeline.config import (
    DEFAULT_PIPELINE_PROFILE_NAME,
    artifact_index_hashes,
    resolve_pipeline_selection,
)
from app.schemas.answer_plan import NormalizedEvidenceItem
from app.services.ai import StubAIService
from app.services.answers import draft_answer_for_row
from app.services.bulk_fill import create_initial_bulk_fill_request, execute_bulk_fill_request
from app.services.cases import create_case_from_uploads, rebuild_case_index_artifacts
from app.services.identity import ensure_local_identity
from app.services.product_truth import ingest_product_truth_file, reimport_product_truth_file
from app.services.seed import import_historical_corpus
from tests.seed_paths import historical_customer_dir, product_truth_path, seed_data_root


def sample_pdf_bytes(repo_root: Path) -> bytes:
    return (
        historical_customer_dir(repo_root, "nordtransit_logistik_ag")
        / "nordtransit_logistik_ag_context_brief.pdf"
    ).read_bytes()


def build_questionnaire_payload(rows: list[tuple[str, str, str]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "QA"
    worksheet["A1"] = "Context"
    worksheet["B1"] = "Question"
    worksheet["C1"] = "Answer"
    for index, (context, question, answer) in enumerate(rows, start=2):
        worksheet.cell(row=index, column=1).value = context
        worksheet.cell(row=index, column=2).value = question
        worksheet.cell(row=index, column=3).value = answer
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_product_truth_source(path: Path, *, count: int) -> Path:
    records = [
        {
            "product_area": "Synthetic Product Truth",
            "title": f"Synthetic truth record {index}",
            "body": (
                f"Synthetic truth record {index} supports workflow routing, reporting, "
                f"integration, and secure document handling for RfX operations."
            ),
            "language": "en",
            "source_file_name": path.name,
            "source_section": f"section-{index}",
            "effective_from": "2026-03-06",
            "effective_to": None,
            "version": f"synthetic-{index}",
        }
        for index in range(1, count + 1)
    ]
    path.write_text(json.dumps(records), encoding="utf-8")
    return path


def create_case(
    session,
    *,
    container,
    repo_root: Path,
    settings,
    pipeline_override: dict[str, object] | None = None,
) -> RfxCase:
    context = ensure_local_identity(session, settings)
    return create_case_from_uploads(
        session,
        storage=container.storage,
        ai_service=StubAIService(),
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        case_name="Pipeline Case",
        client_name="Pipeline Client",
        pdf_file_name="context.pdf",
        pdf_media_type="application/pdf",
        pdf_payload=sample_pdf_bytes(repo_root),
        questionnaire_file_name="qa.xlsx",
        questionnaire_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        questionnaire_payload=build_questionnaire_payload(
            [("Context A", "Question A", ""), ("Context B", "Question B", "")]
        ),
        settings=settings,
        pipeline_override=pipeline_override,
    )


class CapturePlanningAIService(StubAIService):
    def __init__(self) -> None:
        super().__init__()
        self.captured_evidence: list[NormalizedEvidenceItem] = []

    def plan_answer(self, **kwargs):  # type: ignore[override]
        self.captured_evidence = list(kwargs["normalized_evidence"])
        return super().plan_answer(**kwargs)


def planning_invocation_metadata(session, *, answer_version):
    retrieval_run = session.get(RetrievalRun, answer_version.retrieval_run_id)
    assert retrieval_run is not None
    execution_run_id = answer_version.execution_run_id
    assert execution_run_id is not None
    invocations = session.scalars(
        select(ModelInvocation)
        .where(ModelInvocation.execution_run_id == execution_run_id)
        .order_by(ModelInvocation.created_at.asc())
    ).all()
    planning = next(
        item
        for item in invocations
        if item.kind == ModelInvocationKind.ANSWER_GENERATION
        and item.metadata_json.get("prompt_family") == "answer_planning"
    )
    return planning.metadata_json


def test_default_pipeline_profile_is_committed_and_resolved(settings) -> None:
    selection = resolve_pipeline_selection(settings)
    assert selection.profile_name == DEFAULT_PIPELINE_PROFILE_NAME
    assert selection.used_default_profile is True
    assert selection.resolved_pipeline.indexing.current_pdf.chunk_unit == "legacy_char"
    assert selection.resolved_pipeline.indexing.current_pdf.chunk_size == 900
    assert selection.resolved_pipeline.indexing.embedding_model == settings.openai_embedding_model
    assert selection.resolved_pipeline.indexing.embedding_dimensions == settings.openai_embedding_dimensions
    assert settings.openai_response_model == "gpt-5.2"
    assert selection.resolved_pipeline.models.case_profile_extraction.model_id == settings.openai_response_model
    assert selection.resolved_pipeline.models.case_profile_extraction.reasoning_effort == "low"
    assert selection.resolved_pipeline.models.answer_planning.model_id == settings.openai_response_model
    assert selection.resolved_pipeline.models.answer_planning.reasoning_effort == "low"
    assert selection.resolved_pipeline.models.answer_rendering.model_id == settings.openai_response_model
    assert selection.resolved_pipeline.models.answer_rendering.reasoning_effort == "low"
    assert selection.resolved_pipeline.retrieval.candidate_pool.current_case_facts is None
    assert selection.resolved_pipeline.retrieval.candidate_pool.current_pdf_evidence == 24
    assert selection.resolved_pipeline.retrieval.candidate_pool.product_truth == 24
    assert selection.resolved_pipeline.retrieval.candidate_pool.historical_exemplars == 40
    assert selection.resolved_pipeline.retrieval.final_quota.current_case_facts == 3
    assert selection.resolved_pipeline.packing.source_block_order == [
        "current_case_facts",
        "raw_current_pdf",
        "product_truth",
        "historical_exemplars",
    ]


def test_default_pipeline_profile_inherits_model_defaults_from_settings(tmp_path: Path) -> None:
    settings = build_settings(
        env_file=None,
        database_url="sqlite+pysqlite:///:memory:",
        storage_root=tmp_path / "storage",
        openai_response_model="gpt-custom-response",
        openai_embedding_model="text-embedding-custom",
    )
    selection = resolve_pipeline_selection(settings)
    assert selection.resolved_pipeline.indexing.embedding_model == "text-embedding-custom"
    assert selection.resolved_pipeline.indexing.embedding_dimensions == settings.openai_embedding_dimensions
    assert selection.resolved_pipeline.models.case_profile_extraction.model_id == "gpt-custom-response"
    assert selection.resolved_pipeline.models.answer_planning.model_id == "gpt-custom-response"
    assert selection.resolved_pipeline.models.answer_rendering.model_id == "gpt-custom-response"


def test_historical_import_reports_progress(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    messages: list[str] = []
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
        settings=settings,
        progress_callback=messages.append,
    )
    assert any("Historical corpus import started" in message for message in messages)
    assert any("Starting case-profile LLM extraction for historical client" in message for message in messages)
    assert any("Embedded historical questionnaire rows" in message for message in messages)
    assert any("Historical corpus import complete" in message for message in messages)


def test_product_truth_import_reports_progress(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    messages: list[str] = []
    ingest_product_truth_file(
        session,
        ai_service=StubAIService(),
        tenant_id=context.tenant.id,
        path=product_truth_path(repo_root),
        settings=settings,
        progress_callback=messages.append,
    )
    assert any("Product-truth import started" in message for message in messages)
    assert any("Embedded product-truth records:" in message for message in messages)
    assert any("Product-truth import complete" in message for message in messages)


def test_create_case_reports_progress(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    messages: list[str] = []
    create_case_from_uploads(
        session,
        storage=container.storage,
        ai_service=StubAIService(),
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        case_name="Progress Case",
        client_name="Progress Client",
        pdf_file_name="context.pdf",
        pdf_media_type="application/pdf",
        pdf_payload=sample_pdf_bytes(repo_root),
        questionnaire_file_name="qa.xlsx",
        questionnaire_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        questionnaire_payload=build_questionnaire_payload(
            [("Context A", "Question A", ""), ("Context B", "Question B", "")]
        ),
        settings=settings,
        progress_callback=messages.append,
    )
    assert any("Creating live case with" in message for message in messages)
    assert any("Embedded current-PDF chunks for case" in message for message in messages)
    assert any("Starting case-profile LLM extraction for live case" in message for message in messages)
    assert any("Parsed questionnaire workbook for new case" in message for message in messages)
    assert any("Created live case" in message for message in messages)


def test_pipeline_validation_rejects_unknown_and_unsupported_fields(settings) -> None:
    with pytest.raises(ValidationFailure, match="validation failed"):
        resolve_pipeline_selection(
            settings,
            override={"retrieval": {"unknown_field": 1}},
        )
    with pytest.raises(ValidationFailure, match="chunk_unit='token'"):
        resolve_pipeline_selection(
            settings,
            override={"indexing": {"current_pdf": {"chunk_size": 901}}},
        )
    with pytest.raises(ValidationFailure, match="currently supports only 0 or 1"):
        resolve_pipeline_selection(
            settings,
            override={"retrieval": {"broadening": {"max_stages": 2}}},
        )
    with pytest.raises(ValidationFailure, match="fixed dimension 1536"):
        resolve_pipeline_selection(
            settings,
            override={"indexing": {"embedding_dimensions": 3072}},
        )
    with pytest.raises(ValidationFailure, match="conflicts with models.answer_rendering.model_id"):
        resolve_pipeline_selection(
            settings,
            override={
                "generation": {"model_id": "gpt-4.1"},
                "models": {"answer_rendering": {"model_id": "gpt-5.2-preview"}},
            },
        )


def test_stage_model_overrides_resolve_with_new_models_surface(settings) -> None:
    selection = resolve_pipeline_selection(
        settings,
        override={
            "models": {
                "case_profile_extraction": {
                    "model_id": "gpt-5.2",
                    "reasoning_effort": "medium",
                },
                "answer_planning": {
                    "model_id": "gpt-5.2",
                    "reasoning_effort": "medium",
                },
                "answer_rendering": {
                    "model_id": "gpt-5.2",
                    "reasoning_effort": "medium",
                },
            }
        },
    )
    assert selection.resolved_pipeline.models.case_profile_extraction.model_id == "gpt-5.2"
    assert selection.resolved_pipeline.models.case_profile_extraction.reasoning_effort == "medium"
    assert selection.resolved_pipeline.models.answer_planning.model_id == "gpt-5.2"
    assert selection.resolved_pipeline.models.answer_planning.reasoning_effort == "medium"
    assert selection.resolved_pipeline.models.answer_rendering.model_id == "gpt-5.2"
    assert selection.resolved_pipeline.models.answer_rendering.reasoning_effort == "medium"


def test_case_creation_persists_default_pipeline_provenance(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    selection = resolve_pipeline_selection(settings)
    artifact_hashes = artifact_index_hashes(selection)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    session.flush()
    assert case.pipeline_profile_name == DEFAULT_PIPELINE_PROFILE_NAME
    assert case.pipeline_config_hash
    assert case.index_config_hash
    profile = session.scalar(select(CaseProfile).where(CaseProfile.case_id == case.id))
    assert profile is not None
    assert profile.index_config_hash == artifact_hashes.case_profile
    chunks = session.scalars(select(PdfChunk).where(PdfChunk.case_id == case.id)).all()
    assert chunks
    assert all(chunk.index_config_hash == artifact_hashes.current_pdf for chunk in chunks)
    assert all(chunk.chunking_version == "pdf_chunker.v1" for chunk in chunks)


def test_runtime_override_changes_pipeline_hash_without_changing_index_hash(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        settings=settings,
        pipeline_override={
            "retrieval": {"final_quota": {"historical_exemplars": 1}},
            "generation": {"target_answer_words_max": 120},
        },
    )
    retrieval_run = session.get(RetrievalRun, result.answer_version.retrieval_run_id)
    assert retrieval_run is not None
    pipeline = retrieval_run.request_context["pipeline"]
    assert pipeline["config_hash"] != case.pipeline_config_hash
    assert pipeline["index_config_hash"] == case.index_config_hash


def test_runtime_override_preserves_case_pinned_runtime_base(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(
        session,
        container=container,
        repo_root=repo_root,
        settings=settings,
        pipeline_override={"generation": {"target_answer_words_max": 90}},
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        settings=settings,
        pipeline_override={"retrieval": {"final_quota": {"historical_exemplars": 1}}},
    )
    retrieval_run = session.get(RetrievalRun, result.answer_version.retrieval_run_id)
    assert retrieval_run is not None
    pipeline = retrieval_run.request_context["pipeline"]
    resolved = pipeline["resolved_config"]
    assert resolved["generation"]["target_answer_words_max"] == 90
    assert resolved["retrieval"]["final_quota"]["historical_exemplars"] == 1


def test_candidate_pool_overrides_bound_replay_candidate_pools(
    session,
    container,
    repo_root: Path,
    settings,
    tmp_path: Path,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    base_path = seed_data_root(repo_root)
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=base_path,
        settings=settings,
    )
    truth_path = write_product_truth_source(tmp_path / "synthetic_product_truth.json", count=6)
    ingest_product_truth_file(
        session,
        ai_service=StubAIService(),
        tenant_id=context.tenant.id,
        path=truth_path,
        settings=settings,
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    pdf_chunk_count = len(session.scalars(select(PdfChunk).where(PdfChunk.case_id == case.id)).all())
    historical_row_count = len(session.scalars(select(HistoricalQARow)).all())
    product_truth_count = len(session.scalars(select(ProductTruthRecord)).all())
    assert pdf_chunk_count >= 2
    assert historical_row_count >= 2
    assert product_truth_count >= 2

    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        settings=settings,
        pipeline_override={
            "retrieval": {
                "candidate_pool": {
                    "current_pdf_evidence": 2,
                    "product_truth": 2,
                    "historical_exemplars": 2,
                }
            }
        },
    )
    retrieval_run = session.get(RetrievalRun, result.answer_version.retrieval_run_id)
    assert retrieval_run is not None
    execution_run = session.get(ExecutionRun, retrieval_run.execution_run_id)
    assert execution_run is not None
    candidate_pools = execution_run.replay_json["candidate_pools"]
    assert len(candidate_pools["raw_current_pdf"]) == min(2, pdf_chunk_count)
    assert len(candidate_pools["product_truth"]) == min(2, product_truth_count)
    assert len(candidate_pools["historical_exemplar"]) == min(2, historical_row_count)


def test_product_truth_candidate_pool_override_is_single_sourced_from_pipeline_config(
    session,
    container,
    repo_root: Path,
    settings,
    tmp_path: Path,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    truth_path = write_product_truth_source(tmp_path / "synthetic_product_truth.json", count=6)
    ingest_product_truth_file(
        session,
        ai_service=StubAIService(),
        tenant_id=context.tenant.id,
        path=truth_path,
        settings=settings,
    )
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None

    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        settings=settings,
        pipeline_override={"retrieval": {"candidate_pool": {"product_truth": 5}}},
    )
    retrieval_run = session.get(RetrievalRun, result.answer_version.retrieval_run_id)
    assert retrieval_run is not None
    assert (
        retrieval_run.request_context["pipeline"]["resolved_config"]["retrieval"]["candidate_pool"][
            "product_truth"
        ]
        == 5
    )
    execution_run = session.get(ExecutionRun, retrieval_run.execution_run_id)
    assert execution_run is not None
    candidate_pools = execution_run.replay_json["candidate_pools"]
    assert len(candidate_pools["product_truth"]) == 5


def test_packing_override_reorders_planning_evidence_and_changes_runtime_hash(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    ai_service = CapturePlanningAIService()
    source_block_order = [
        "raw_current_pdf",
        "current_case_facts",
        "product_truth",
        "historical_exemplars",
    ]
    result = draft_answer_for_row(
        session,
        ai_service=ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        settings=settings,
        pipeline_override={"packing": {"source_block_order": source_block_order}},
    )
    retrieval_run = session.get(RetrievalRun, result.answer_version.retrieval_run_id)
    assert retrieval_run is not None
    pipeline = retrieval_run.request_context["pipeline"]
    assert pipeline["config_hash"] != case.pipeline_config_hash
    assert pipeline["index_config_hash"] == case.index_config_hash
    captured_layers = [item.layer for item in ai_service.captured_evidence]
    assert captured_layers
    assert "raw_current_pdf" in captured_layers
    assert "current_case_facts" in captured_layers
    assert captured_layers[0] == "raw_current_pdf"
    first_current_case_index = captured_layers.index("current_case_facts")
    assert all(layer == "raw_current_pdf" for layer in captured_layers[:first_current_case_index])
    metadata = planning_invocation_metadata(session, answer_version=result.answer_version)
    packing = metadata["packing"]
    assert packing["order_strategy"] == "source_block_order"
    assert packing["source_block_order"] == source_block_order
    assert packing["output_layers"] == captured_layers
    assert packing["truncated"] is False


def test_packing_budget_fails_closed_when_first_item_cannot_fit(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    with pytest.raises(
        ValidationFailure,
        match="too small to fit even the first packed evidence item",
    ):
        draft_answer_for_row(
            session,
            ai_service=CapturePlanningAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
            settings=settings,
            pipeline_override={
                "models": {"answer_planning": {"model_id": "gpt-4o-mini"}},
                "packing": {"max_context_tokens": 1},
            },
        )


def test_incompatible_index_override_fails_loudly(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    with pytest.raises(ValidationFailure, match="current-PDF index hash"):
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
            settings=settings,
            pipeline_override={
                "indexing": {
                    "current_pdf": {
                        "chunk_unit": "token",
                        "chunk_size": 120,
                        "chunk_overlap": 20,
                    }
                }
            },
        )


def test_bulk_fill_rejects_incompatible_index_override_before_queueing(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    assert questionnaire is not None
    with pytest.raises(ValidationFailure, match="current-PDF index hash"):
        create_initial_bulk_fill_request(
            session,
            case=case,
            questionnaire=questionnaire,
            user_id=context.user.id,
            note="bad index override",
            settings=settings,
            pipeline_override={
                "indexing": {
                    "current_pdf": {
                        "chunk_unit": "token",
                        "chunk_size": 120,
                        "chunk_overlap": 20,
                    }
                }
            },
        )


def test_bulk_fill_pins_pipeline_config_consistently(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    assert questionnaire is not None
    request = create_initial_bulk_fill_request(
        session,
        case=case,
        questionnaire=questionnaire,
        user_id=context.user.id,
        note="pipeline bulk fill",
        settings=settings,
        pipeline_override={"generation": {"target_answer_words_max": 80}},
    )
    request_id = request.id
    expected_hash = request.config_json["pipeline_config_hash"]
    session.commit()

    execute_bulk_fill_request(
        container,
        request_id=request_id,
        runner_id="pipeline-test-runner",
        execution_mode="inline-test",
    )

    with container.session_factory() as check_session:
        request = check_session.get(BulkFillRequest, request_id)
        assert request is not None
        assert request.config_json["pipeline_config_hash"] == expected_hash
        retrieval_runs = check_session.scalars(
            select(RetrievalRun).where(RetrievalRun.case_id == case.id)
        ).all()
        assert retrieval_runs
        assert all(
            run.request_context["pipeline"]["config_hash"] == expected_hash
            for run in retrieval_runs
        )


def test_pipeline_default_endpoint_exposes_default_profile(client) -> None:
    response = client.get("/api/pipeline-config/default")
    assert response.status_code == 200
    body = response.json()
    assert body["profile_name"] == "default"
    assert body["config"]["pipeline_version"] == "rfx_pipeline.v1"
    assert body["artifact_index_hashes"]["current_pdf"]
    assert body["config_schema"]


def test_non_default_token_chunking_changes_pdf_chunk_artifacts(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    default_case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    default_chunks = session.scalars(
        select(PdfChunk)
        .where(PdfChunk.case_id == default_case.id)
        .order_by(PdfChunk.page_number.asc(), PdfChunk.chunk_index.asc())
    ).all()
    assert default_chunks
    token_case = create_case(
        session,
        container=container,
        repo_root=repo_root,
        settings=settings,
        pipeline_override={
            "indexing": {
                "current_pdf": {
                    "chunk_unit": "token",
                    "chunk_size": 120,
                    "chunk_overlap": 20,
                }
            }
        },
    )
    token_chunks = session.scalars(
        select(PdfChunk)
        .where(PdfChunk.case_id == token_case.id)
        .order_by(PdfChunk.page_number.asc(), PdfChunk.chunk_index.asc())
    ).all()
    assert token_chunks
    assert all(chunk.chunking_version == "pdf_chunker.v2" for chunk in token_chunks)
    assert len(token_chunks) != len(default_chunks) or [
        chunk.content for chunk in token_chunks
    ] != [chunk.content for chunk in default_chunks]


def test_contextualized_chunking_changes_chunk_content_and_hash(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    default_case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    default_chunk = session.scalar(
        select(PdfChunk)
        .where(PdfChunk.case_id == default_case.id)
        .order_by(PdfChunk.page_number.asc(), PdfChunk.chunk_index.asc())
    )
    assert default_chunk is not None
    contextualized_case = create_case(
        session,
        container=container,
        repo_root=repo_root,
        settings=settings,
        pipeline_override={"indexing": {"current_pdf": {"contextualize_chunks": True}}},
    )
    contextualized_chunk = session.scalar(
        select(PdfChunk)
        .where(PdfChunk.case_id == contextualized_case.id)
        .order_by(PdfChunk.page_number.asc(), PdfChunk.chunk_index.asc())
    )
    assert contextualized_chunk is not None
    assert contextualized_chunk.chunking_version == "pdf_chunker.v2"
    assert contextualized_chunk.content.startswith("Case: Pipeline Case")
    assert contextualized_chunk.content != default_chunk.content
    assert contextualized_chunk.chunk_hash != default_chunk.chunk_hash


def test_token_chunking_with_unknown_embedding_model_fails_loudly(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    with pytest.raises(ValidationFailure, match="does not support embedding model"):
        create_case(
            session,
            container=container,
            repo_root=repo_root,
            settings=settings,
            pipeline_override={
                "indexing": {
                    "embedding_model": "unsupported-tokenizer-model",
                    "current_pdf": {
                        "chunk_unit": "token",
                        "chunk_size": 120,
                        "chunk_overlap": 20,
                    },
                }
            },
        )


def test_live_case_rebuild_under_new_index_config_updates_lineage_and_unblocks_requests(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    token_override = {
        "indexing": {
            "current_pdf": {
                "chunk_unit": "token",
                "chunk_size": 120,
                "chunk_overlap": 20,
            }
        }
    }
    with pytest.raises(ValidationFailure, match="current-PDF index hash"):
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
            settings=settings,
            pipeline_override=token_override,
        )
    rebuild_case_index_artifacts(
        session,
        storage=container.storage,
        ai_service=StubAIService(),
        case=case,
        settings=settings,
        pipeline_override=token_override,
    )
    session.flush()
    selection = resolve_pipeline_selection(
        settings,
        pinned_config=case.pipeline_config_json,
        pinned_profile_name=case.pipeline_profile_name,
    )
    artifact_hashes = artifact_index_hashes(selection)
    rebuilt_chunks = session.scalars(select(PdfChunk).where(PdfChunk.case_id == case.id)).all()
    assert rebuilt_chunks
    assert all(chunk.index_config_hash == artifact_hashes.current_pdf for chunk in rebuilt_chunks)
    assert all(chunk.chunking_version == "pdf_chunker.v2" for chunk in rebuilt_chunks)
    result = draft_answer_for_row(
        session,
        ai_service=StubAIService(),
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        settings=settings,
        pipeline_override=token_override,
    )
    retrieval_run = session.get(RetrievalRun, result.answer_version.retrieval_run_id)
    assert retrieval_run is not None
    assert retrieval_run.request_context["pipeline"]["index_config_hash"] == case.index_config_hash


def test_live_case_rebuild_marks_replaced_builds_and_links_replacement_chain(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    original_current_pdf_build_id = case.current_pdf_build_id
    original_case_profile_build_id = case.case_profile_build_id
    assert original_current_pdf_build_id is not None
    assert original_case_profile_build_id is not None
    token_override = {
        "indexing": {
            "current_pdf": {
                "chunk_unit": "token",
                "chunk_size": 120,
                "chunk_overlap": 20,
            }
        }
    }
    rebuild_case_index_artifacts(
        session,
        storage=container.storage,
        ai_service=StubAIService(),
        case=case,
        settings=settings,
        pipeline_override=token_override,
    )
    session.flush()
    assert case.current_pdf_build_id is not None
    assert case.case_profile_build_id is not None
    assert case.current_pdf_build_id != original_current_pdf_build_id
    assert case.case_profile_build_id != original_case_profile_build_id

    old_current_pdf_build = session.get(ArtifactBuild, original_current_pdf_build_id)
    old_case_profile_build = session.get(ArtifactBuild, original_case_profile_build_id)
    new_current_pdf_build = session.get(ArtifactBuild, case.current_pdf_build_id)
    new_case_profile_build = session.get(ArtifactBuild, case.case_profile_build_id)
    assert old_current_pdf_build is not None
    assert old_case_profile_build is not None
    assert new_current_pdf_build is not None
    assert new_case_profile_build is not None
    assert old_current_pdf_build.status == ArtifactBuildStatus.REPLACED
    assert old_case_profile_build.status == ArtifactBuildStatus.REPLACED
    assert new_current_pdf_build.status == ArtifactBuildStatus.ACTIVE
    assert new_case_profile_build.status == ArtifactBuildStatus.ACTIVE
    assert new_current_pdf_build.replaced_build_id == old_current_pdf_build.id
    assert new_case_profile_build.replaced_build_id == old_case_profile_build.id


def test_historical_reimport_under_new_index_config_updates_lineage(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    base_path = seed_data_root(repo_root)
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=base_path,
        settings=settings,
    )
    original_package = session.scalar(select(HistoricalClientPackage))
    assert original_package is not None
    original_hash = original_package.index_config_hash
    override = {"indexing": {"historical": {"signature_mode": "summary_only"}}}
    import_historical_corpus(
        session,
        ai_service=StubAIService(),
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=base_path,
        settings=settings,
        pipeline_override=override,
    )
    new_package = session.scalar(select(HistoricalClientPackage))
    assert new_package is not None
    assert new_package.index_config_hash != original_hash


def test_product_truth_reimport_under_new_index_config_updates_lineage(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    path = product_truth_path(repo_root)
    ingest_product_truth_file(
        session,
        ai_service=StubAIService(),
        tenant_id=context.tenant.id,
        path=path,
        settings=settings,
    )
    original_record = session.scalar(select(ProductTruthRecord))
    assert original_record is not None
    original_hash = original_record.index_config_hash
    reimport_product_truth_file(
        session,
        ai_service=StubAIService(),
        tenant_id=context.tenant.id,
        path=path,
        settings=settings,
        pipeline_override={"indexing": {"embedding_model": "text-embedding-3-large"}},
    )
    new_record = session.scalar(select(ProductTruthRecord))
    assert new_record is not None
    assert new_record.index_config_hash != original_hash


def test_mixed_pdf_chunk_lineage_fails_loudly_after_rebuild(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    row = session.scalar(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    )
    assert row is not None
    token_override = {
        "indexing": {
            "current_pdf": {
                "chunk_unit": "token",
                "chunk_size": 120,
                "chunk_overlap": 20,
            }
        }
    }
    rebuild_case_index_artifacts(
        session,
        storage=container.storage,
        ai_service=StubAIService(),
        case=case,
        settings=settings,
        pipeline_override=token_override,
    )
    bad_chunk = session.scalar(select(PdfChunk).where(PdfChunk.case_id == case.id))
    assert bad_chunk is not None
    bad_chunk.index_config_hash = "bogus-hash"
    session.flush()
    with pytest.raises(ValidationFailure, match="current-PDF index hash"):
        draft_answer_for_row(
            session,
            ai_service=StubAIService(),
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
            settings=settings,
            pipeline_override=token_override,
        )
