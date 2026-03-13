from __future__ import annotations

import json
from dataclasses import replace
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from uuid import UUID

import pytest
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import Settings, build_settings
from app.exceptions import ValidationFailure
from app.models.entities import (
    AnswerVersion,
    CaseProfile,
    CaseProfileItem,
    HistoricalCaseProfile,
    HistoricalCaseProfileItem,
    HistoricalClientPackage,
    HistoricalDataset,
    HistoricalQARow,
    HistoricalWorkbook,
    ModelInvocation,
    PdfChunk,
    ProductTruthChunk,
    ProductTruthRecord,
    Questionnaire,
    QuestionnaireRow,
    RfxCase,
    Upload,
)
from app.models.enums import ApprovalStatus, ModelInvocationKind, UploadKind
from app.pipeline.config import (
    artifact_index_hashes,
    historical_index_payload,
    product_truth_index_payload,
    resolve_pipeline_selection,
)
from app.prompts.answer_planning import ANSWER_PLANNING_SYSTEM_PROMPT
from app.schemas.answer_plan import AnswerPlan
from app.schemas.case_profile_extraction import CaseProfileExtractionOutput
from app.services.ai import StubAIService
from app.services.answers import draft_answer_for_row
from app.services.case_profiles import generate_case_profile_document
from app.services.cases import create_case_from_uploads
from app.services.hashing import sha256_hex, sha256_text
from app.services.identity import ensure_local_identity
from app.services.pdf_chunks import current_pdf_chunking_version
from tests.seed_paths import historical_customer_dir

HISTORICAL_SCHEMA_VERSION = "historical_qa_workbook.v1"


def sample_pdf_bytes(repo_root: Path) -> bytes:
    return (
        historical_customer_dir(repo_root, "nordtransit_logistik_ag")
        / "nordtransit_logistik_ag_context_brief.pdf"
    ).read_bytes()


def build_questionnaire_payload(rows: list[tuple[str, str, str]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "QA"
    worksheet["A1"] = "Context"
    worksheet["B1"] = "Question"
    worksheet["C1"] = "Answer"
    for index, (context, question, answer) in enumerate(rows, start=2):
        worksheet.cell(row=index, column=1).value = context
        worksheet.cell(row=index, column=2).value = question
        worksheet.cell(row=index, column=3).value = answer
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_manual_case_with_profile(
    session: Session,
    *,
    tenant_id,
    user_id,
    ai_service: StubAIService,
    row_question: str,
    row_context: str,
    case_fact_items: list[tuple[str, str]] | None = None,
) -> tuple[RfxCase, QuestionnaireRow]:
    pipeline = resolve_pipeline_selection(build_settings(env_file=None))
    artifact_hashes = artifact_index_hashes(pipeline)
    case = RfxCase(
        tenant_id=tenant_id,
        created_by_user_id=user_id,
        name="NordTransit Prompt Case",
        client_name="NordTransit Logistik AG",
        language="de",
        pipeline_profile_name=pipeline.profile_name,
        pipeline_config_json=pipeline.resolved_config,
        pipeline_config_hash=pipeline.config_hash,
        index_config_hash=pipeline.index_config_hash,
    )
    session.add(case)
    session.flush()
    pdf_upload = Upload(
        tenant_id=tenant_id,
        case_id=case.id,
        kind=UploadKind.CASE_PDF,
        original_file_name="context.pdf",
        media_type="application/pdf",
        object_key=f"tests/{case.id}/case-pdf",
        file_hash="case-pdf-hash",
        size_bytes=1,
        payload=b"x",
    )
    questionnaire_upload = Upload(
        tenant_id=tenant_id,
        case_id=case.id,
        kind=UploadKind.QUESTIONNAIRE_XLSX,
        original_file_name="questionnaire.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        object_key=f"tests/{case.id}/questionnaire",
        file_hash="questionnaire-hash",
        size_bytes=1,
        payload=b"x",
    )
    session.add_all([pdf_upload, questionnaire_upload])
    session.flush()
    profile = CaseProfile(
        tenant_id=tenant_id,
        case_id=case.id,
        source_pdf_upload_id=pdf_upload.id,
        schema_version="rfx_case_profile.v3",
        prompt_set_version="rfx_case_profile_prompt_set.v3",
        model="stub-ai-service",
        summary="NordTransit seeks a staged multi-depot rollout with depot-level workflow support.",
        source_file_name="context.pdf",
        source_file_hash="case-pdf-hash",
        language="de",
        pipeline_profile_name=pipeline.profile_name,
        index_config_json={"embedding_model": "stub-ai-service"},
        index_config_hash=artifact_hashes.case_profile,
        generated_at=datetime.now(UTC),
        document={"case_id": str(case.id)},
    )
    session.add(profile)
    session.flush()
    fact_items = case_fact_items or [
        (
            "initiative_scope",
            "NordTransit requests a phased introduction across multiple depots with depot-by-depot sequencing.",
        ),
        (
            "business_capabilities_in_scope",
            "The requested scope centers on transport execution and depot workflow support.",
        ),
    ]
    for position, (item_id, answer) in enumerate(fact_items, start=1):
        session.add(
            CaseProfileItem(
                tenant_id=tenant_id,
                case_profile_id=profile.id,
                case_id=case.id,
                analysis_item_id=item_id,
                position=position,
                prompt=item_id,
                answer=answer,
                confidence="high",
                citations=["Page 1"],
                normalized_text=answer,
                embedding=ai_service.embed_text(answer),
                )
            )
    chunk_text = (
        "NordTransit requires customer-specific grounding for depot workflows and phased rollout across depots."
    )
    session.add(
        PdfChunk(
            tenant_id=tenant_id,
            case_id=case.id,
            upload_id=pdf_upload.id,
            page_number=1,
            chunk_index=0,
            start_offset=0,
            end_offset=len(chunk_text),
            chunking_version=current_pdf_chunking_version(pipeline),
            embedding_model=pipeline.resolved_pipeline.indexing.embedding_model or "stub-ai-service",
            index_config_hash=artifact_hashes.current_pdf,
            artifact_build_id=None,
            chunk_hash=sha256_text(chunk_text),
            content=chunk_text,
            embedding=ai_service.embed_text(chunk_text),
        )
    )
    questionnaire = Questionnaire(
        tenant_id=tenant_id,
        case_id=case.id,
        upload_id=questionnaire_upload.id,
        source_file_name="questionnaire.xlsx",
        source_sheet_name="QA",
        file_hash="questionnaire-hash",
        schema_version="questionnaire_workbook.v1",
    )
    session.add(questionnaire)
    session.flush()
    row = QuestionnaireRow(
        tenant_id=tenant_id,
        case_id=case.id,
        questionnaire_id=questionnaire.id,
        source_sheet_name="QA",
        source_row_number=2,
        source_row_id="questionnaire.xlsx:QA:2",
        context_raw=row_context,
        question_raw=row_question,
        answer_raw="",
        normalized_text=f"{row_context}\n{row_question}",
    )
    session.add(row)
    session.flush()
    return case, row


def create_historical_package(
    session: Session,
    *,
    tenant_id,
    ai_service: StubAIService,
    client_slug: str,
    client_name: str,
    row_context: str,
    row_question: str,
    row_answer: str,
) -> HistoricalQARow:
    pipeline = resolve_pipeline_selection(build_settings(env_file=None))
    artifact_hashes = artifact_index_hashes(pipeline)
    dataset = HistoricalDataset(
        tenant_id=tenant_id,
        name="Prompt Regression Dataset",
        slug=f"prompt-regression-{client_slug}",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(dataset)
    session.flush()
    workbook = HistoricalWorkbook(
        tenant_id=tenant_id,
        dataset_id=dataset.id,
        client_name=client_name,
        language="de",
        source_file_name=f"{client_slug}.xlsx",
        source_sheet_name="QA",
        file_hash=f"{client_slug}-xlsx-hash",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        approval_status=ApprovalStatus.APPROVED,
    )
    session.add(workbook)
    session.flush()
    package = HistoricalClientPackage(
        tenant_id=tenant_id,
        dataset_id=dataset.id,
        workbook_id=workbook.id,
        client_slug=client_slug,
        client_name=client_name,
        language="de",
        source_pdf_file_name=f"{client_slug}.pdf",
        source_pdf_file_hash=f"{client_slug}-pdf-hash",
        pdf_object_key=f"tests/historical/{client_slug}.pdf",
        pdf_size_bytes=1,
        pipeline_profile_name=pipeline.profile_name,
        index_config_json=historical_index_payload(pipeline.resolved_pipeline),
        index_config_hash=artifact_hashes.historical,
    )
    session.add(package)
    session.flush()
    profile = HistoricalCaseProfile(
        tenant_id=tenant_id,
        client_package_id=package.id,
        schema_version="rfx_case_profile.v3",
        prompt_set_version="rfx_case_profile_prompt_set.v3",
        model="stub-ai-service",
        summary=f"{client_name} operates outside the NordTransit logistics domain.",
        source_file_name=f"{client_slug}.pdf",
        source_file_hash=f"{client_slug}-pdf-hash",
        language="de",
        generated_at=datetime.now(UTC),
        signature_version="case_signature.v1",
        signature_embedding_model="stub-ai-service",
        signature_fields_json={"summary": client_name},
        signature_text=client_name,
        signature_embedding=ai_service.embed_text(client_name),
        document={"client_name": client_name},
    )
    session.add(profile)
    session.flush()
    session.add(
        HistoricalCaseProfileItem(
            tenant_id=tenant_id,
            historical_case_profile_id=profile.id,
            analysis_item_id="strategic_objectives",
            position=1,
            prompt="strategic_objectives",
            answer=client_name,
            confidence="high",
            citations=["citation_unavailable"],
            normalized_text=client_name,
            embedding=ai_service.embed_text(client_name),
        )
    )
    row = HistoricalQARow(
        tenant_id=tenant_id,
        dataset_id=dataset.id,
        workbook_id=workbook.id,
        client_package_id=package.id,
        historical_case_profile_id=profile.id,
        client_name=client_name,
        source_file_name=f"{client_slug}.xlsx",
        source_sheet_name="QA",
        source_row_number=2,
        source_row_id=f"{client_slug}:QA:2",
        language="de",
        language_confidence=1.0,
        approval_status=ApprovalStatus.APPROVED,
        context_raw=row_context,
        question_raw=row_question,
        answer_raw=row_answer,
        normalized_text=f"{row_context}\n{row_question}",
        file_hash=f"{client_slug}-xlsx-hash",
        schema_version=HISTORICAL_SCHEMA_VERSION,
        embedding=ai_service.embed_text(f"{row_context}\n{row_question}"),
    )
    session.add(row)
    session.flush()
    return row


def add_product_truth_record(
    session: Session,
    *,
    tenant_id,
    ai_service: StubAIService,
    title: str,
    body: str,
) -> None:
    pipeline = resolve_pipeline_selection(build_settings(env_file=None))
    artifact_hashes = artifact_index_hashes(pipeline)
    record = ProductTruthRecord(
        tenant_id=tenant_id,
        product_area="core_suite",
        title=title,
        body=body,
        language="de",
        source_file_name="product_truth.json",
        source_section=title,
        effective_from=date(2025, 1, 1),
        effective_to=None,
        version="v1",
        pipeline_profile_name=pipeline.profile_name,
        index_config_json=product_truth_index_payload(pipeline.resolved_pipeline),
        index_config_hash=artifact_hashes.product_truth,
        artifact_build_id=None,
        approval_status=ApprovalStatus.APPROVED,
        file_hash=sha256_hex(body.encode("utf-8")),
    )
    session.add(record)
    session.flush()
    session.add(
        ProductTruthChunk(
            tenant_id=tenant_id,
            truth_record_id=record.id,
            chunk_index=0,
            content=body,
            language="de",
            file_hash=record.file_hash,
            approval_status=ApprovalStatus.APPROVED,
            embedding=ai_service.embed_text(body),
        )
    )
    session.flush()


def planning_invocation_payload(session: Session, *, answer_version: AnswerVersion) -> dict[str, object]:
    rendering_run_id = answer_version.execution_run_id
    assert rendering_run_id is not None
    invocations = session.scalars(
        select(ModelInvocation)
        .where(ModelInvocation.execution_run_id == rendering_run_id)
        .order_by(ModelInvocation.created_at.asc())
    ).all()
    planning = next(
        item
        for item in invocations
        if item.kind == ModelInvocationKind.ANSWER_GENERATION
        and item.metadata_json.get("prompt_family") == "answer_planning"
    )
    assert planning.response_payload_text is not None
    return json.loads(planning.response_payload_text)


def test_answer_planning_prompt_has_split_claim_sections_and_examples() -> None:
    prompt = ANSWER_PLANNING_SYSTEM_PROMPT
    for header in (
        "ROLE AND OUTPUT",
        "AUTHORITY MODEL",
        "CLAIM SPLITTING",
        "QUESTION INTENT POLICY",
        "CLAIM POLICY",
        "UNKNOWN POLICY",
        "RULES",
    ):
        assert header in prompt
    assert "\n\n\n\nAUTHORITY MODEL" in prompt
    assert "\n\n\n\nCLAIM SPLITTING" in prompt
    assert "Valid example:" in prompt
    assert "Invalid example:" in prompt
    assert "emit separate claims" in prompt
    assert "Never mix `current_case_facts` or `raw_current_pdf` support ids into `product_mapping`, `integration`, `security`, or `reporting` claims." in prompt


def test_rollout_answer_regression_is_customer_facing(
    session: Session,
    container,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Der Rollout soll über mehrere Depots mit begrenzter Betriebsunterbrechung erfolgen.",
        row_question="Wie unterstützen Sie eine schrittweise Einführung über mehrere Depots?",
    )
    add_product_truth_record(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        title="BluePeak Core Platform",
        body="BluePeak Core Platform supports staged activation of depot workflows and transport execution capabilities.",
    )
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        client_slug="historical-rollout",
        client_name="Asteron Insurance SE",
        row_context="Mehrstufige Einführung über Regionen.",
        row_question="Wie wurde der Rollout strukturiert?",
        row_answer="1. Current case facts 2. Product truth 3. Historical exemplars. BluePeak Pulse delivered a successful rollout.",
    )
    result = draft_answer_for_row(
        session,
        ai_service=ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Bitte formuliere eine präzise Kundenantwort.",
        thread=None,
    )
    answer = result.answer_version.answer_text
    assert answer.strip()
    assert "Current case facts" not in answer
    assert "Product truth" not in answer
    assert "Historical exemplars" not in answer
    assert not answer.lstrip().startswith("1.")
    assert "wir unterstützen sie gerne" not in answer.lower()


def test_product_fit_answer_regression_stays_grounded(
    session: Session,
    container,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Gefragt ist die Abdeckung des Kernprozesses für Transport- und Depotabläufe.",
        row_question="Welche Ihrer Produkte decken den Kern des angefragten Prozesses inklusive Reporting und Integrationen ab?",
    )
    add_product_truth_record(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        title="BluePeak Flow",
        body="BluePeak Flow supports transport execution and depot workflow orchestration. The current approved truth does not confirm reporting modules or a full integration catalog.",
    )
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        client_slug="historical-product-fit",
        client_name="Asteron Insurance SE",
        row_context="Versicherungsprozesse, nicht Logistik.",
        row_question="Welche Lösung wurde eingesetzt?",
        row_answer="BluePeak Pulse showed successful rollout in logistics and covered reporting.",
    )
    result = draft_answer_for_row(
        session,
        ai_service=ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Bitte beantworte die Frage kurz und belastbar.",
        thread=None,
    )
    answer = result.answer_version.answer_text
    plan = AnswerPlan.model_validate(planning_invocation_payload(session, answer_version=result.answer_version))
    assert "BluePeak Pulse" not in answer
    assert "Asteron" not in answer
    assert "successful rollout" not in answer.lower()
    assert "Offen bleibt" in answer
    assert all(
        claim.authority_layer == "product_truth"
        for claim in plan.supported_claims
        if claim.claim_kind in {"product_mapping", "integration", "security", "reporting"}
    )


def test_customer_facing_render_mode_never_leaks_internal_headings(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Der Fokus liegt auf der Produktabdeckung des Kernprozesses.",
        row_question="Welche Ihrer Produkte decken den Kern des angefragten Prozesses ab?",
    )
    add_product_truth_record(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        title="BluePeak Flow",
        body="BluePeak Flow supports the requested transport execution scope.",
    )
    result = draft_answer_for_row(
        session,
        ai_service=ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Antwort bitte direkt formulieren.",
        thread=None,
    )
    answer = result.answer_version.answer_text
    for heading in (
        "Current case facts",
        "Product truth",
        "Historical exemplars",
        "Aktuelle Fallfakten",
        "Produktwahrheit",
        "Historische Beispiele",
    ):
        assert heading not in answer


def test_integration_claims_require_product_truth_support(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="SAP, Azure AD und Microsoft 365 sind im Kundenkontext vorhanden.",
        row_question="Wie integriert sich Ihre Lösung mit SAP und Azure AD?",
    )
    result = draft_answer_for_row(
        session,
        ai_service=ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Bitte beantworte die Integrationsfrage.",
        thread=None,
    )
    plan = AnswerPlan.model_validate(planning_invocation_payload(session, answer_version=result.answer_version))
    assert not any(claim.claim_kind == "integration" for claim in plan.supported_claims)
    assert any("Integrationsansatz" in unknown.topic or "integration approach" in unknown.topic for unknown in plan.unknowns)


def test_security_claims_require_product_truth_support(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Der Kunde fordert EU-Hosting, Audit-Trail und rollenbasierte Zugriffe.",
        row_question="Wie adressieren Sie EU-Hosting, Auditierbarkeit und rollenbasierte Zugriffskontrolle?",
    )
    result = draft_answer_for_row(
        session,
        ai_service=ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Bitte beantworte die Sicherheitsfrage.",
        thread=None,
    )
    plan = AnswerPlan.model_validate(planning_invocation_payload(session, answer_version=result.answer_version))
    assert not any(claim.claim_kind == "security" for claim in plan.supported_claims)
    assert any("Compliance" in unknown.topic or "hosting" in unknown.topic or "compliance" in unknown.topic.lower() for unknown in plan.unknowns)


def test_historical_exemplars_remain_non_factual_for_product_fit(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Gefragt ist die Produktabdeckung des Kernprozesses.",
        row_question="Welche Ihrer Produkte decken den Kern des angefragten Prozesses ab?",
    )
    add_product_truth_record(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        title="BluePeak Flow",
        body="BluePeak Flow supports the requested process scope.",
    )
    create_historical_package(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        client_slug="historical-style",
        client_name="Asteron Insurance SE",
        row_context="Nicht logistischer Kontext.",
        row_question="Welche Lösung wurde eingesetzt?",
        row_answer="BluePeak Pulse covered the process and delivered a successful rollout.",
    )
    result = draft_answer_for_row(
        session,
        ai_service=ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Bitte beantworte die Produktfrage.",
        thread=None,
    )
    plan_payload = planning_invocation_payload(session, answer_version=result.answer_version)
    plan = AnswerPlan.model_validate(plan_payload)
    assert plan.primary_intent == "product_fit"
    assert all(claim.authority_layer != "historical_exemplar" for claim in plan.supported_claims)
    assert all(isinstance(hint, str) for hint in plan.historical_style_guidance)
    assert not any("source_id" in hint for hint in plan.historical_style_guidance)


def test_invalid_case_profile_extraction_output_fails_loudly(
    session: Session,
    container,
    repo_root: Path,
    settings: Settings,
) -> None:
    class InvalidExtractionAIService(StubAIService):
        def generate_case_profile(self, **kwargs):  # type: ignore[override]
            result = super().generate_case_profile(**kwargs)
            bad_output = CaseProfileExtractionOutput.model_construct(
                schema_version=result.structured_output.schema_version,
                analysis_items=result.structured_output.analysis_items[:-1],
                summary=result.structured_output.summary,
            )
            return replace(result, structured_output=bad_output)

    context = ensure_local_identity(session, settings)
    with pytest.raises(ValidationFailure, match="extraction output invalid"):
        create_case_from_uploads(
            session,
            storage=container.storage,
            ai_service=InvalidExtractionAIService(),
            tenant_id=context.tenant.id,
            user_id=context.user.id,
            case_name="Invalid Extraction",
            client_name="NordTransit Logistik AG",
            pdf_file_name="context.pdf",
            pdf_media_type="application/pdf",
            pdf_payload=sample_pdf_bytes(repo_root),
            questionnaire_file_name="qa.xlsx",
            questionnaire_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            questionnaire_payload=build_questionnaire_payload([("Kontext", "Frage", "")]),
        )


def test_invalid_answer_plan_fails_loudly(
    session: Session,
    settings: Settings,
) -> None:
    class InvalidAnswerPlanAIService(StubAIService):
        def plan_answer(self, **kwargs):  # type: ignore[override]
            result = super().plan_answer(**kwargs)
            bad_plan = AnswerPlan.model_construct(
                **{
                    **result.answer_plan.model_dump(mode="python"),
                    "supported_claims": [
                        {
                            "claim": "Unsupported claim",
                            "claim_kind": "product_mapping",
                            "authority_layer": "product_truth",
                            "support_ids": ["missing-support-id"],
                        }
                    ],
                }
            )
            return replace(
                result,
                answer_plan=bad_plan,
                response_payload=bad_plan.model_dump(mode="json"),
            )

    context = ensure_local_identity(session, settings)
    ai_service = InvalidAnswerPlanAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Gefragt ist die Produktabdeckung des Kernprozesses.",
        row_question="Welche Ihrer Produkte decken den Kern des angefragten Prozesses ab?",
    )
    add_product_truth_record(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        title="BluePeak Flow",
        body="BluePeak Flow supports the requested process scope.",
    )
    with pytest.raises(ValidationFailure, match="unknown support id"):
        draft_answer_for_row(
            session,
            ai_service=ai_service,
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Bitte beantworte die Frage.",
            thread=None,
        )


def test_mixed_authority_product_claim_fails_loudly(
    session: Session,
    settings: Settings,
) -> None:
    class MixedAuthorityAnswerPlanAIService(StubAIService):
        def plan_answer(self, **kwargs):  # type: ignore[override]
            result = super().plan_answer(**kwargs)
            bad_plan = AnswerPlan.model_construct(
                **{
                    **result.answer_plan.model_dump(mode="python"),
                    "supported_claims": [
                        {
                            "claim": "BluePeak Flow fits the requested scope.",
                            "claim_kind": "product_mapping",
                            "authority_layer": "product_truth",
                            "support_ids": ["CF1", "PT1"],
                        }
                    ],
                }
            )
            return replace(
                result,
                answer_plan=bad_plan,
                response_payload=bad_plan.model_dump(mode="json"),
            )

    context = ensure_local_identity(session, settings)
    ai_service = MixedAuthorityAnswerPlanAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="The client asks which product fits the requested scope.",
        row_question="Which products fit the requested scope?",
    )
    add_product_truth_record(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        title="BluePeak Flow",
        body="BluePeak Flow supports configurable workflow automation.",
    )
    with pytest.raises(ValidationFailure, match="incompatible evidence layers"):
        draft_answer_for_row(
            session,
            ai_service=ai_service,
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Please answer the product-fit question.",
            thread=None,
        )


def test_extraction_output_stays_ledger_like_under_stub_contract(settings: Settings) -> None:
    result = generate_case_profile_document(
        ai_service=StubAIService(),
        pipeline=resolve_pipeline_selection(settings),
        case_id=UUID("00000000-0000-0000-0000-000000000001"),
        source_file_name="context.pdf",
        source_file_hash="hash",
        client_name="Ledger Test Client",
        language="de",
        page_text=["Kurzbeschreibung des Projekts mit Scope, Zeitplan und Integrationslandschaft."],
    )
    first_item = result.structured_output.analysis_items[0]
    assert first_item.id == "strategic_objectives"
    assert "strategic_objectives" not in first_item.answer
    assert len(first_item.answer.split()) <= 20


def test_compound_question_keeps_secondary_intent(
    session: Session,
    settings: Settings,
) -> None:
    context = ensure_local_identity(session, settings)
    ai_service = StubAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Der Rollout soll über mehrere Depots phasenweise erfolgen.",
        row_question="Welche Ihrer Produkte decken den Kern des angefragten Prozesses ab und wie unterstützen Sie eine schrittweise Einführung über mehrere Depots?",
    )
    add_product_truth_record(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        title="BluePeak Flow",
        body="BluePeak Flow supports the requested process scope and can be deployed in phased rollout programs.",
    )
    result = draft_answer_for_row(
        session,
        ai_service=ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Bitte beantworte die zusammengesetzte Frage.",
        thread=None,
    )
    plan = AnswerPlan.model_validate(planning_invocation_payload(session, answer_version=result.answer_version))
    assert plan.primary_intent == "product_fit"
    assert "rollout_approach" in plan.secondary_intents


def test_unknown_topics_must_be_plain_language(
    session: Session,
    settings: Settings,
) -> None:
    class SnakeCaseUnknownAIService(StubAIService):
        def plan_answer(self, **kwargs):  # type: ignore[override]
            result = super().plan_answer(**kwargs)
            bad_plan = AnswerPlan.model_construct(
                **{
                    **result.answer_plan.model_dump(mode="python"),
                    "unknowns": [
                        {
                            "topic": "reporting_support",
                            "reason": "Missing proof.",
                            "materiality": "material",
                        }
                    ],
                }
            )
            return replace(
                result,
                answer_plan=bad_plan,
                response_payload=bad_plan.model_dump(mode="json"),
            )

    context = ensure_local_identity(session, settings)
    ai_service = SnakeCaseUnknownAIService()
    case, row = create_manual_case_with_profile(
        session,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        ai_service=ai_service,
        row_context="Reporting darf nur genannt werden, wenn es belegt ist.",
        row_question="Welche Ihrer Produkte decken den Kern des angefragten Prozesses ab?",
    )
    add_product_truth_record(
        session,
        tenant_id=context.tenant.id,
        ai_service=ai_service,
        title="BluePeak Flow",
        body="BluePeak Flow supports the requested process scope.",
    )
    with pytest.raises(ValidationFailure, match="snake_case"):
        draft_answer_for_row(
            session,
            ai_service=ai_service,
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Bitte beantworte die Frage.",
            thread=None,
        )
