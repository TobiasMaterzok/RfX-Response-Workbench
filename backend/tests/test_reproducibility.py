from __future__ import annotations

from io import BytesIO
from pathlib import Path
from uuid import UUID

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.exceptions import ValidationFailure
from app.models.entities import (
    ArtifactBuild,
    ExecutionRun,
    ModelInvocation,
    ProductTruthRecord,
    RepoSnapshot,
    RfxCase,
    RuntimeSnapshot,
    SourceManifest,
)
from app.models.enums import (
    ArtifactBuildKind,
    ExecutionRunKind,
    ExportMode,
    QuestionnaireRowStatus,
    ReproducibilityLevel,
    ReproducibilityMode,
    SourceManifestKind,
)
from app.services.answers import draft_answer_for_row
from app.services.bulk_fill import approve_answer_version
from app.services.cases import create_case_from_uploads
from app.services.exports import export_questionnaire
from app.services.identity import ensure_local_identity
from app.services.product_truth import reimport_product_truth_file
from app.services.reproducibility import build_execution_run_manifest, start_repro_run
from app.services.seed import import_historical_corpus
from tests.seed_paths import historical_customer_dir, product_truth_path, seed_data_root


def sample_pdf_bytes(repo_root: Path) -> bytes:
    return (
        historical_customer_dir(repo_root, "nordtransit_logistik_ag")
        / "nordtransit_logistik_ag_context_brief.pdf"
    ).read_bytes()


def build_questionnaire_payload(rows: list[tuple[str, str, str]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "QA"
    worksheet["A1"] = "Context"
    worksheet["B1"] = "Question"
    worksheet["C1"] = "Answer"
    for index, (context, question, answer) in enumerate(rows, start=2):
        worksheet.cell(row=index, column=1).value = context
        worksheet.cell(row=index, column=2).value = question
        worksheet.cell(row=index, column=3).value = answer
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def enable_strict_eval_runtime(monkeypatch) -> None:
    from app.services import reproducibility

    monkeypatch.setattr(reproducibility, "_alembic_head", lambda session: "test-head")


def create_case(
    session,
    *,
    container,
    repo_root: Path,
    settings,
    reproducibility_mode: ReproducibilityMode = ReproducibilityMode.BEST_EFFORT,
) -> RfxCase:
    context = ensure_local_identity(session, settings)
    return create_case_from_uploads(
        session,
        storage=container.storage,
        ai_service=container.ai_service,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        case_name="Repro Case",
        client_name="Repro Client",
        pdf_file_name="context.pdf",
        pdf_media_type="application/pdf",
        pdf_payload=sample_pdf_bytes(repo_root),
        questionnaire_file_name="qa.xlsx",
        questionnaire_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        questionnaire_payload=build_questionnaire_payload([("Context A", "Question A", "")]),
        settings=settings,
        reproducibility_mode=reproducibility_mode,
    )


def test_case_creation_captures_snapshots_runs_builds_and_manifest(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    assert case.creation_run_id is not None
    assert case.current_pdf_build_id is not None
    assert case.case_profile_build_id is not None
    assert session.scalar(select(RepoSnapshot)) is not None
    assert session.scalar(select(RuntimeSnapshot)) is not None
    manifest = session.scalar(select(SourceManifest).where(SourceManifest.case_id == case.id))
    assert manifest is not None
    assert manifest.kind == SourceManifestKind.LIVE_CASE_INPUT
    run = session.get(ExecutionRun, case.creation_run_id)
    assert run is not None
    assert run.kind == ExecutionRunKind.LIVE_CASE_CREATE
    builds = session.scalars(
        select(ArtifactBuild).where(
            ArtifactBuild.case_id == case.id,
            ArtifactBuild.kind.in_({ArtifactBuildKind.CURRENT_PDF, ArtifactBuildKind.CASE_PROFILE}),
        )
    ).all()
    assert len(builds) == 2


def test_strict_eval_case_creation_records_case_profile_lineage(
    session,
    container,
    repo_root: Path,
    settings,
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    case = create_case(
        session,
        container=container,
        repo_root=repo_root,
        settings=settings,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    run = session.get(ExecutionRun, case.creation_run_id)
    assert run is not None
    assert run.reproducibility_level == ReproducibilityLevel.OPERATIONALLY_REPLAYABLE
    manifest = build_execution_run_manifest(session, run_id=run.id)
    invocation = next(
        item for item in manifest["model_invocations"] if item["kind"] == "case_profile_extraction"
    )
    assert manifest["runtime_snapshot"]["env_fingerprint"]["storage_root"] == str(
        settings.storage_root.resolve()
    )
    assert invocation["metadata"]["resolved_prompt_hash"]


def test_strict_eval_requires_backend_lockfile(monkeypatch, session, settings) -> None:
    from app.services import reproducibility

    monkeypatch.setattr(reproducibility, "BACKEND_LOCKFILE", Path("/missing/backend.lock"))
    with pytest.raises(ValidationFailure, match="backend lockfile"):
        start_repro_run(
            session,
            storage=None,
            settings=settings,
            kind=ExecutionRunKind.EXPORT,
            mode=ReproducibilityMode.STRICT_EVAL,
        )


def test_best_effort_repo_snapshot_capture_failure_is_explicit(monkeypatch, session, settings) -> None:
    from app.services import reproducibility

    monkeypatch.setattr(reproducibility, "_git_command", lambda *args: (_ for _ in ()).throw(RuntimeError("git unavailable")))
    repro = start_repro_run(
        session,
        storage=None,
        settings=settings,
        kind=ExecutionRunKind.EXPORT,
        mode=ReproducibilityMode.BEST_EFFORT,
    )
    snapshot = session.get(RepoSnapshot, repro.repo_snapshot.id)
    assert snapshot is not None
    assert snapshot.git_commit_sha == "capture_failed"
    assert snapshot.git_dirty is True
    assert snapshot.git_diff_text is not None
    assert "repo_snapshot_capture_failed" in snapshot.git_diff_text


def test_historical_import_creates_run_build_and_manifest(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    dataset = import_historical_corpus(
        session,
        ai_service=container.ai_service,
        storage=container.storage,
        tenant_id=context.tenant.id,
        base_path=seed_data_root(repo_root),
        settings=settings,
    )
    assert dataset.creation_run_id is not None
    assert dataset.artifact_build_id is not None
    manifest = session.scalar(
        select(SourceManifest).where(
            SourceManifest.kind == SourceManifestKind.HISTORICAL_IMPORT_SOURCE
        )
    )
    assert manifest is not None
    run = session.get(ExecutionRun, dataset.creation_run_id)
    assert run is not None
    assert run.kind == ExecutionRunKind.HISTORICAL_IMPORT


def test_product_truth_reimport_creates_run_and_build(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    records = reimport_product_truth_file(
        session,
        storage=container.storage,
        ai_service=container.ai_service,
        tenant_id=context.tenant.id,
        path=product_truth_path(repo_root),
        settings=settings,
    )
    assert records
    assert all(record.artifact_build_id is not None for record in records)
    run = session.scalar(
        select(ExecutionRun).where(ExecutionRun.kind == ExecutionRunKind.PRODUCT_TRUTH_REIMPORT)
    )
    assert run is not None


def test_product_truth_api_import_creates_run_and_build(
    client,
    auth_headers: dict[str, str],
    session,
) -> None:
    response = client.post(
        "/api/product-truth/import",
        headers=auth_headers,
        json={
            "product_area": "api_area",
            "title": "API truth",
            "body": "API-imported truth record.",
            "language": "en",
            "source_file_name": "api.json",
            "source_section": "api_section",
            "effective_from": "2026-03-08",
            "effective_to": None,
            "version": "api-v1",
        },
    )
    assert response.status_code == 200
    record = session.get(ProductTruthRecord, UUID(response.json()["truth_record_id"]))
    assert record is not None
    assert record.artifact_build_id is not None
    run = session.scalar(
        select(ExecutionRun).where(ExecutionRun.kind == ExecutionRunKind.PRODUCT_TRUTH_IMPORT)
    )
    assert run is not None


def test_product_truth_api_strict_eval_forbids_additive_import(
    client,
    auth_headers: dict[str, str],
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    baseline = {
        "product_area": "api_area",
        "title": "API truth",
        "body": "API-imported truth record.",
        "language": "en",
        "source_file_name": "api.json",
        "source_section": "api_section",
        "effective_from": "2026-03-08",
        "effective_to": None,
        "version": "api-v1",
    }
    initial = client.post(
        "/api/product-truth/import",
        headers=auth_headers,
        json=baseline,
    )
    assert initial.status_code == 200
    strict_eval = client.post(
        "/api/product-truth/import",
        headers=auth_headers,
        json={
            **baseline,
            "title": "API truth strict",
            "version": "api-v2",
            "reproducibility_mode": "strict_eval",
        },
    )
    assert strict_eval.status_code == 422
    assert "strict_eval forbids additive product-truth import" in strict_eval.json()["detail"]


def test_draft_records_correct_model_invocation_and_retrieval_replay_state(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    from app.models.entities import AnswerVersion, QuestionnaireRow, RetrievalRun

    question_row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert question_row is not None
    result = draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=question_row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        pipeline_override={"generation": {"model_id": "gpt-4.1"}},
    )
    answer = session.get(AnswerVersion, result.answer_version.id)
    retrieval = session.get(RetrievalRun, result.answer_version.retrieval_run_id)
    assert answer is not None
    assert retrieval is not None
    assert answer.execution_run_id is not None
    assert answer.model_invocation_id is not None
    assert answer.model == "gpt-4.1"
    invocation = session.get(ModelInvocation, answer.model_invocation_id)
    assert invocation is not None
    assert invocation.metadata_json["prompt_family"] == "answer_rendering"
    assert invocation.metadata_json["resolved_prompt_hash"]
    planning_invocation = next(
        (
            item
            for item in session.scalars(
                select(ModelInvocation).where(
                    ModelInvocation.execution_run_id == answer.execution_run_id
                )
            ).all()
            if item.metadata_json.get("prompt_family") == "answer_planning"
        ),
        None,
    )
    assert planning_invocation is not None
    assert retrieval.execution_run_id is not None
    retrieval_run = session.get(ExecutionRun, retrieval.execution_run_id)
    assert retrieval_run is not None
    assert retrieval_run.replay_json is not None
    assert "candidate_pools" in retrieval_run.replay_json
    assert "request_date" in retrieval_run.replay_json


def test_strict_eval_retrieval_is_operationally_replayable_and_records_query_embeddings(
    session,
    container,
    repo_root: Path,
    settings,
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    context = ensure_local_identity(session, settings)
    case = create_case(
        session,
        container=container,
        repo_root=repo_root,
        settings=settings,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    from app.models.entities import QuestionnaireRow, RetrievalRun

    row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert row is not None
    result = draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    retrieval = session.get(RetrievalRun, result.answer_version.retrieval_run_id)
    assert retrieval is not None
    retrieval_run = session.get(ExecutionRun, retrieval.execution_run_id)
    assert retrieval_run is not None
    assert retrieval_run.reproducibility_level == ReproducibilityLevel.OPERATIONALLY_REPLAYABLE
    assert retrieval_run.replay_json is not None
    assert "query_embeddings" in retrieval_run.replay_json
    assert retrieval_run.replay_json["query_embeddings"]


def test_strict_eval_rejects_legacy_case_builds_for_draft(
    session,
    container,
    repo_root: Path,
    settings,
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    case.current_pdf_build_id = None
    session.flush()
    from app.models.entities import QuestionnaireRow

    row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert row is not None
    with pytest.raises(ValidationFailure, match="missing required current-case build lineage"):
        draft_answer_for_row(
            session,
            ai_service=container.ai_service,
            case=case,
            row=row,
            user_id=context.user.id,
            user_message="Draft the answer.",
            thread=None,
            reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
        )


def test_export_creates_run_manifest_and_stable_canonical_export_manifest(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    from app.models.entities import Questionnaire, QuestionnaireRow, Upload
    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert questionnaire is not None
    assert row is not None
    draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None
    export_job = export_questionnaire(
        session,
        storage=container.storage,
        settings=settings,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
        user_id=context.user.id,
    )
    assert export_job.execution_run_id is not None
    manifest_one = build_execution_run_manifest(session, run_id=export_job.execution_run_id)
    manifest_two = build_execution_run_manifest(session, run_id=export_job.execution_run_id)
    assert manifest_one["manifest_hash"] == manifest_two["manifest_hash"]


def test_strict_eval_export_is_deterministic_non_llm(
    session,
    container,
    repo_root: Path,
    settings,
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    context = ensure_local_identity(session, settings)
    case = create_case(
        session,
        container=container,
        repo_root=repo_root,
        settings=settings,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    from app.models.entities import Questionnaire, QuestionnaireRow, Upload

    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert questionnaire is not None
    assert row is not None
    draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None
    export_job = export_questionnaire(
        session,
        storage=container.storage,
        settings=settings,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
        user_id=context.user.id,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    run = session.get(ExecutionRun, export_job.execution_run_id)
    assert run is not None
    assert run.reproducibility_level == ReproducibilityLevel.DETERMINISTIC_NON_LLM
    assert isinstance(export_job.metadata_json.get("csv_upload_id"), str)
    assert isinstance(export_job.metadata_json.get("zip_upload_id"), str)


def test_strict_eval_approved_only_export_accepts_placeholder_rows(
    session,
    container,
    repo_root: Path,
    settings,
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    context = ensure_local_identity(session, settings)
    case = create_case_from_uploads(
        session,
        storage=container.storage,
        ai_service=container.ai_service,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        case_name="Strict Eval Approved Export",
        client_name="Repro Client",
        pdf_file_name="context.pdf",
        pdf_media_type="application/pdf",
        pdf_payload=sample_pdf_bytes(repo_root),
        questionnaire_file_name="qa.xlsx",
        questionnaire_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        questionnaire_payload=build_questionnaire_payload(
            [("Context A", "Question A", ""), ("Context B", "Question B", "")]
        ),
        settings=settings,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    from app.models.entities import Questionnaire, QuestionnaireRow, Upload

    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    assert questionnaire is not None
    rows = session.scalars(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    ).all()
    assert len(rows) == 2
    approved_draft = draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=rows[0],
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    approve_answer_version(
        session,
        case=case,
        row=rows[0],
        answer_version_id=approved_draft.answer_version.id,
    )
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None

    export_job = export_questionnaire(
        session,
        storage=container.storage,
        settings=settings,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.APPROVED_ONLY,
        user_id=context.user.id,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    run = session.get(ExecutionRun, export_job.execution_run_id)
    assert run is not None
    assert run.reproducibility_level == ReproducibilityLevel.DETERMINISTIC_NON_LLM
    build_execution_run_manifest(session, run_id=export_job.execution_run_id)
    source_manifest = session.get(SourceManifest, run.source_manifest_id)
    assert source_manifest is not None
    row_selection = source_manifest.manifest_json["row_selection"]
    assert row_selection[0]["selection_kind"] == "approved_answer"
    assert row_selection[0]["answer_version_id"] == str(approved_draft.answer_version.id)
    assert row_selection[1]["selection_kind"] == "status_placeholder"
    assert row_selection[1]["answer_version_id"] is None
    assert row_selection[1]["placeholder_text"] == (
        "No approved answer exported due to status: not started."
    )
    assert isinstance(export_job.metadata_json.get("csv_upload_id"), str)
    assert isinstance(export_job.metadata_json.get("zip_upload_id"), str)


def test_strict_eval_latest_available_export_accepts_placeholder_rows(
    session,
    container,
    repo_root: Path,
    settings,
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    context = ensure_local_identity(session, settings)
    case = create_case_from_uploads(
        session,
        storage=container.storage,
        ai_service=container.ai_service,
        tenant_id=context.tenant.id,
        user_id=context.user.id,
        case_name="Strict Eval Latest Export",
        client_name="Repro Client",
        pdf_file_name="context.pdf",
        pdf_media_type="application/pdf",
        pdf_payload=sample_pdf_bytes(repo_root),
        questionnaire_file_name="qa.xlsx",
        questionnaire_media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        questionnaire_payload=build_questionnaire_payload(
            [("Context A", "Question A", ""), ("Context B", "Question B", "")]
        ),
        settings=settings,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    from app.models.entities import Questionnaire, QuestionnaireRow, Upload

    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    assert questionnaire is not None
    rows = session.scalars(
        select(QuestionnaireRow)
        .where(QuestionnaireRow.case_id == case.id)
        .order_by(QuestionnaireRow.source_row_number.asc())
    ).all()
    assert len(rows) == 2
    draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=rows[0],
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    rows[1].review_status = QuestionnaireRowStatus.FAILED
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None

    export_job = export_questionnaire(
        session,
        storage=container.storage,
        settings=settings,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
        user_id=context.user.id,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    build_execution_run_manifest(session, run_id=export_job.execution_run_id)
    run = session.get(ExecutionRun, export_job.execution_run_id)
    assert run is not None
    source_manifest = session.get(SourceManifest, run.source_manifest_id)
    assert source_manifest is not None
    row_selection = source_manifest.manifest_json["row_selection"]
    assert row_selection[0]["selection_kind"] == "approved_answer"
    assert row_selection[1]["selection_kind"] == "status_placeholder"
    assert row_selection[1]["answer_version_id"] is None
    assert row_selection[1]["placeholder_text"] == (
        "No latest answer exported due to status: failed."
    )
    assert isinstance(export_job.metadata_json.get("csv_upload_id"), str)
    assert isinstance(export_job.metadata_json.get("zip_upload_id"), str)


def test_strict_eval_export_rejects_answers_without_full_lineage(
    session,
    container,
    repo_root: Path,
    settings,
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    from app.models.entities import Questionnaire, QuestionnaireRow, Upload

    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert questionnaire is not None
    assert row is not None
    draft = draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    draft.answer_version.model_invocation_id = None
    session.flush()
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None
    with pytest.raises(ValidationFailure, match="missing strict-eval lineage"):
        export_questionnaire(
            session,
            storage=container.storage,
            settings=settings,
            questionnaire=questionnaire,
            upload=upload,
            mode=ExportMode.LATEST_AVAILABLE,
            user_id=context.user.id,
            reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
        )


def test_strict_eval_manifest_detects_cross_table_export_drift(
    session,
    container,
    repo_root: Path,
    settings,
    monkeypatch,
) -> None:
    enable_strict_eval_runtime(monkeypatch)
    context = ensure_local_identity(session, settings)
    case = create_case(
        session,
        container=container,
        repo_root=repo_root,
        settings=settings,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    from app.models.entities import AnswerVersion, Questionnaire, QuestionnaireRow, Upload

    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert questionnaire is not None
    assert row is not None
    draft = draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None
    export_job = export_questionnaire(
        session,
        storage=container.storage,
        settings=settings,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
        user_id=context.user.id,
        reproducibility_mode=ReproducibilityMode.STRICT_EVAL,
    )
    answer = session.get(AnswerVersion, draft.answer_version.id)
    assert answer is not None
    answer.execution_run_id = None
    session.flush()
    with pytest.raises(ValidationFailure, match="without full lineage"):
        build_execution_run_manifest(session, run_id=export_job.execution_run_id)


def test_export_manifest_detects_missing_csv_upload_lineage(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    from app.models.entities import Questionnaire, QuestionnaireRow, Upload

    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert questionnaire is not None
    assert row is not None
    draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None
    export_job = export_questionnaire(
        session,
        storage=container.storage,
        settings=settings,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
        user_id=context.user.id,
    )
    run = session.get(ExecutionRun, export_job.execution_run_id)
    assert run is not None
    export_job.metadata_json.pop("csv_upload_id", None)
    session.flush()
    with pytest.raises(ValidationFailure, match="missing CSV upload lineage"):
        build_execution_run_manifest(session, run_id=export_job.execution_run_id)


def test_export_manifest_detects_missing_zip_upload_lineage(
    session,
    container,
    repo_root: Path,
    settings,
) -> None:
    context = ensure_local_identity(session, settings)
    case = create_case(session, container=container, repo_root=repo_root, settings=settings)
    from app.models.entities import Questionnaire, QuestionnaireRow, Upload

    questionnaire = session.scalar(select(Questionnaire).where(Questionnaire.case_id == case.id))
    row = session.scalar(select(QuestionnaireRow).where(QuestionnaireRow.case_id == case.id))
    assert questionnaire is not None
    assert row is not None
    draft_answer_for_row(
        session,
        ai_service=container.ai_service,
        case=case,
        row=row,
        user_id=context.user.id,
        user_message="Draft the answer.",
        thread=None,
    )
    upload = session.get(Upload, questionnaire.upload_id)
    assert upload is not None
    export_job = export_questionnaire(
        session,
        storage=container.storage,
        settings=settings,
        questionnaire=questionnaire,
        upload=upload,
        mode=ExportMode.LATEST_AVAILABLE,
        user_id=context.user.id,
    )
    run = session.get(ExecutionRun, export_job.execution_run_id)
    assert run is not None
    export_job.metadata_json.pop("zip_upload_id", None)
    session.flush()
    with pytest.raises(ValidationFailure, match="missing ZIP upload lineage"):
        build_execution_run_manifest(session, run_id=export_job.execution_run_id)
